import os
import re
import sys
import time
import json
import random
import shutil
import logging
import threading
import atexit
import subprocess
import platform
from datetime import datetime, timedelta
from collections import deque
from concurrent.futures import ProcessPoolExecutor, as_completed

import pandas as pd
import requests
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, WebDriverException,
    StaleElementReferenceException
)
from webdriver_manager.chrome import ChromeDriverManager

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

try:
    from urllib3.exceptions import ProtocolError, MaxRetryError, ReadTimeoutError
except Exception:
    class ProtocolError(Exception):
        ...
    class MaxRetryError(Exception):
        ...
    class ReadTimeoutError(Exception):
        ...

try:
    import psutil
except Exception:
    psutil = None


def get_executable_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


exec_dir = get_executable_dir()
os.chdir(exec_dir)

images_dir = os.path.join(exec_dir, 'images')
os.makedirs(images_dir, exist_ok=True)

BASE_USER_DATA_ROOT = os.path.join(exec_dir, 'chrome_profile', 'User Data')
ISOLATED_USER_DATA_ROOT = os.path.join(exec_dir, 'chrome_profile', 'isolated')
os.makedirs(BASE_USER_DATA_ROOT, exist_ok=True)
os.makedirs(ISOLATED_USER_DATA_ROOT, exist_ok=True)

logging.basicConfig(
    filename=os.path.join(exec_dir, 'tiktok.log'),
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.INFO
)


def _rotate_logs(max_size_mb: int = 10, keep_days: int = 14):
    try:
        path = os.path.join(exec_dir, 'tiktok.log')
        if os.path.exists(path):
            size_mb = os.path.getsize(path) / (1024 * 1024)
            if size_mb > max_size_mb:
                arch = os.path.join(exec_dir, f'tiktok_{datetime.now():%Y%m%d_%H%M%S}.log')
                shutil.copy(path, arch)
                open(path, 'w', encoding='utf-8').close()
    except Exception as e:
        logging.warning(f'log rotation error: {e}')


_rotate_logs()

_DEFAULT_CFG = {
    "WEBDRIVER_CMD_TIMEOUT_SEC": 600,
    "SCRIPT_TIMEOUT_SEC": 180,
    "EXPLICIT_WAIT_SEC": 30,
    "PER_SONG_TIME_BUDGET_SEC": 300,
    "PARALLEL_ENABLED": False,
    "PARALLEL_WORKERS": 2,
    "CHROME_PROFILE_POOL": ["Default", "Profile 1"],
    "SONG_INTERVAL_MIN_SEC": 8,
    "SONG_INTERVAL_MAX_SEC": 15,
    "ERROR_PAGE_EXTRA_WAIT_SEC": 10,
    "MAX_CONSECUTIVE_ERRORS": 3,
    "LONG_WAIT_AFTER_ERRORS_SEC": 60,
    # 【追加】ログイン確認関連
    "CHECK_LOGIN_BEFORE_START": True,
    "LOGIN_WAIT_TIMEOUT_SEC": 120,
}


def _load_runtime_config() -> dict:
    cfg = dict(_DEFAULT_CFG)
    cands = []
    envp = os.environ.get("TIKTOK_CONFIG")
    if envp:
        cands.append(envp)
    cands.append(os.path.join(exec_dir, "config.json"))
    cands.append("config.json")
    for p in cands:
        try:
            if p and os.path.isfile(p):
                with open(p, "r", encoding="utf-8") as f:
                    user = json.load(f) or {}
                cfg.update(user)
                logging.info(f'Config loaded: {os.path.abspath(p)}')
                break
        except Exception as e:
            logging.warning(f'Config load failed ({p}): {e}')
    for k in ["WEBDRIVER_CMD_TIMEOUT_SEC", "SCRIPT_TIMEOUT_SEC", "EXPLICIT_WAIT_SEC",
              "PER_SONG_TIME_BUDGET_SEC", "PARALLEL_WORKERS",
              "SONG_INTERVAL_MIN_SEC", "SONG_INTERVAL_MAX_SEC",
              "ERROR_PAGE_EXTRA_WAIT_SEC", "MAX_CONSECUTIVE_ERRORS",
              "LONG_WAIT_AFTER_ERRORS_SEC", "LOGIN_WAIT_TIMEOUT_SEC"]:
        try:
            cfg[k] = int(cfg.get(k, _DEFAULT_CFG.get(k, 0)))
        except Exception:
            cfg[k] = _DEFAULT_CFG.get(k, 0)
    try:
        cfg["PARALLEL_WORKERS"] = max(1, min(int(cfg.get("PARALLEL_WORKERS", 2)), 2))
    except Exception:
        cfg["PARALLEL_WORKERS"] = 2
    return cfg


CFG = _load_runtime_config()

stop_flag = threading.Event()
external_stop = None


def is_stopped() -> bool:
    try:
        if stop_flag.is_set():
            return True
        if external_stop is not None and external_stop.is_set():
            return True
    except Exception:
        pass
    return False


def signal_global_stop():
    try:
        stop_flag.set()
        if external_stop is not None:
            external_stop.set()
    except Exception:
        pass


def is_service_like_session() -> bool:
    try:
        if os.environ.get("TIKTOK_SCHEDULER", "") == "1":
            return True
        sname = os.environ.get("SESSIONNAME", "")
    except Exception:
        sname = ""
    us = (sname or "").upper()
    if not us:
        return True
    if us.startswith("SERVICE"):
        return True
    if us.startswith("DWM-"):
        return True
    return False


class TextHandler(logging.Handler):
    def __init__(self, text_widget, log_display_var):
        super().__init__()
        self.text_widget = text_widget
        self.log_display_var = log_display_var

    def emit(self, record):
        if self.log_display_var.get():
            msg = self.format(record)

            def append():
                self.text_widget.configure(state='normal')
                self.text_widget.insert(tk.END, msg + '\n')
                self.text_widget.configure(state='disabled')
                self.text_widget.yview(tk.END)

            self.text_widget.after(0, append)


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def sanitize_filename(name: str) -> str:
    name = name if isinstance(name, str) else 'unknown_song'
    return re.sub(r'[\\/:*?"<>|\r\n]+', '_', name)


def parse_number(text: str):
    if not text:
        return None
    t = text.strip().replace(',', '')
    mult = 1
    if t[-1:] in ('K', 'k'):
        mult = 1_000
        t = t[:-1]
    elif t[-1:] in ('M', 'm'):
        mult = 1_000_000
        t = t[:-1]
    try:
        return int(float(t) * mult)
    except Exception:
        return None


def _is_invalid_session_error(exc: Exception) -> bool:
    try:
        msg = (str(exc) or '').lower()
    except Exception:
        msg = ''
    keywords = [
        'invalid session id',
        'chrome not reachable',
        'session deleted',
        'no such window',
    ]
    return any(k in msg for k in keywords)


class DriverInitFatalError(Exception):
    pass


def resolve_settings_path(user_specified: str | None) -> str:
    if user_specified and os.path.exists(user_specified):
        return os.path.abspath(user_specified)
    cands = [
        os.path.join(exec_dir, 'initial_settings.xlsx'),
        os.path.join(exec_dir, 'input', 'initial_settings.xlsx'),
    ]
    for p in cands:
        if os.path.exists(p):
            return os.path.abspath(p)
    if (len(sys.argv) == 1 or sys.argv[1] != 'run') and not is_service_like_session():
        p = filedialog.askopenfilename(
            title='initial_settings.xlsx を選択してください',
            filetypes=[('Excel Files', '*.xlsx *.xlsm *.xltx *.xltm')]
        )
        if p:
            return os.path.abspath(p)
    raise FileNotFoundError('初期設定Excelファイルが見つかりませんでした。')


def read_initial_settings(settings_path: str | None = None):
    try:
        resolved = resolve_settings_path(settings_path)
        logging.info(f'初期設定Excel: {resolved}')
        df_settings = pd.read_excel(resolved, sheet_name='初期設定', header=None)
        save_path = df_settings.iloc[0, 1]
        max_items = df_settings.iloc[2, 1]
        max_items = int(max_items) if pd.notna(max_items) else 0

        if isinstance(save_path, str):
            ensure_dir(save_path)

        df_urls = pd.read_excel(resolved, sheet_name='取得楽曲URL設定')
        empty_idx = df_urls[(df_urls['曲名'].isnull()) & (df_urls['URL'].isnull())].index
        if not empty_idx.empty:
            df_urls = df_urls.iloc[:empty_idx[0]]
        df_urls = df_urls.dropna(subset=['曲名', 'URL'])
        df_urls['曲名'] = df_urls['曲名'].astype(str)
        df_urls['URL'] = df_urls['URL'].astype(str) \
            .str.replace(r'^\s*ttps://', 'https://', regex=True) \
            .str.replace(r'[\r\n]+', '', regex=True)
        return save_path, max_items, df_urls[['曲名', 'URL']].values.tolist()
    except Exception as e:
        logging.error(f'初期設定の読み込み中にエラー: {e}')
        try:
            messagebox.showerror('エラー', f'初期設定の読み込み中にエラーが発生しました: {e}')
        except Exception:
            pass
        return None, None, None


webdriver_processes: list[int] = []
launched_user_dirs: list[str] = []


def _cmdline_contains_any(proc, substrings: list[str]) -> bool:
    try:
        cmd = ' '.join(proc.cmdline())
        return any(s for s in substrings if s and s in cmd)
    except Exception:
        return False


def kill_chrome_processes():
    if not psutil:
        logging.info('psutil 未導入のため、明示プロセス終了はスキップします。')
        return
    for pid in list(webdriver_processes):
        try:
            p = psutil.Process(pid)
            p.terminate()
        except psutil.NoSuchProcess:
            pass
        except Exception as e:
            logging.warning(f'プロセス {pid} 終了時エラー: {e}')
        finally:
            try:
                webdriver_processes.remove(pid)
            except Exception:
                pass

    try:
        targets = []
        marker_paths = set(launched_user_dirs + [ISOLATED_USER_DATA_ROOT])
        for p in psutil.process_iter(['name', 'cmdline']):
            nm = (p.info.get('name') or '').lower()
            if nm not in ('chrome.exe', 'chromedriver.exe'):
                continue
            if _cmdline_contains_any(p, [str(m) for m in marker_paths]) or _cmdline_contains_any(p, ['--remote-debugging-port=0']):
                targets.append(p)
        for p in targets:
            try:
                p.terminate()
            except Exception:
                pass
        time.sleep(0.8)
        for p in targets:
            try:
                if p.is_running():
                    p.kill()
            except Exception:
                pass
    except Exception as e:
        logging.warning(f'Chrome/Driver 掃除処理で例外: {e}')


def force_kill_all_processes():
    try:
        system = platform.system().lower()
    except Exception:
        system = ''
    if system != 'windows':
        return
    targets = ('tiktok.exe', 'chromedriver.exe', 'chrome.exe')
    for name in targets:
        try:
            subprocess.run(
                ['taskkill', '/f', '/t', '/im', name],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False
            )
        except Exception as e:
            logging.info(f'taskkill 失敗 ({name}): {e}')


atexit.register(kill_chrome_processes)


def _copytree_safe(src: str, dst: str):
    def _ignore(_dir, names):
        ignore_patterns = [
            'Singleton*', 'LOCK', 'Lockfile', 'Visited Links', 'Current Tabs',
            'Current Session', 'Last Tabs', 'Last Session', 'Crashpad', 'Code Cache'
        ]
        res = []
        for n in names:
            for pat in ignore_patterns:
                if re.fullmatch(pat.replace('*', '.*'), n):
                    res.append(n)
                    break
        return set(res)

    shutil.copytree(src, dst, dirs_exist_ok=True, ignore=_ignore)


def seed_isolated_profile(profile_name: str, isolated_ud_dir: str):
    try:
        marker = os.path.join(isolated_ud_dir, '.SEEDED_OK')
        dst_profile = os.path.join(isolated_ud_dir, 'Default')
        if os.path.exists(marker) and os.path.isdir(dst_profile):
            return

        src_profile = os.path.join(BASE_USER_DATA_ROOT, profile_name)
        if not os.path.isdir(src_profile):
            logging.warning(f'シード元プロファイルが見つかりません: {src_profile}（未ログインで起動します）')
            ensure_dir(dst_profile)
        else:
            logging.info(f'プロファイルシード開始: {profile_name} -> {dst_profile}')
            ensure_dir(isolated_ud_dir)
            _copytree_safe(src_profile, dst_profile)
            logging.info('プロファイルシード完了')

        local_state_src = os.path.join(BASE_USER_DATA_ROOT, 'Local State')
        if os.path.isfile(local_state_src):
            try:
                shutil.copy2(local_state_src, os.path.join(isolated_ud_dir, 'Local State'))
            except Exception:
                pass

        with open(marker, 'w', encoding='utf-8') as f:
            f.write('ok')
    except Exception as e:
        logging.warning(f'プロファイルシード中に例外: {e}')


def init_driver(headless: bool = False, per_song_timeout: int = 300,
                for_function1: bool = False, profile_name: str | None = None,
                chromedriver_path: str | None = None,
                _is_retry: bool = False):
    if is_service_like_session() and not headless:
        logging.info('サービス/スケジューラっぽいセッションのため、Chrome を headless モードで起動します。')
        headless = True

    chrome_options = Options()

    prefs = {
        "profile.managed_default_content_settings.images": 1 if for_function1 else 2,
        "disk-cache-size": 4096
    }
    chrome_options.add_experimental_option("prefs", prefs)
    if headless:
        chrome_options.add_argument("--headless=new")

    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--no-first-run")
    chrome_options.add_argument("--no-default-browser-check")
    chrome_options.add_argument("--disable-background-networking")
    chrome_options.add_argument("--disable-renderer-backgrounding")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-features=Translate,OptimizationHints")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--remote-debugging-port=0")
    chrome_options.add_argument("--mute-audio")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.set_capability("pageLoadStrategy", "none")

    if profile_name:
        ud_dir = os.path.abspath(os.path.join(ISOLATED_USER_DATA_ROOT, f'{sanitize_filename(profile_name)}'))
        ensure_dir(ud_dir)
        seed_isolated_profile(profile_name, ud_dir)
        chrome_options.add_argument(f'--user-data-dir={ud_dir}')
        chrome_options.add_argument('--profile-directory=Default')
        logging.info(f'Using isolated Chrome user-data-dir: {ud_dir}')
    else:
        ud_dir = BASE_USER_DATA_ROOT
        ensure_dir(ud_dir)
        chrome_options.add_argument(f'--user-data-dir={ud_dir}')
        chrome_options.add_argument('--profile-directory=Default')
        logging.info(f'Using shared Chrome user-data-dir: {ud_dir}')

    try:
        if chromedriver_path and os.path.exists(chromedriver_path):
            service = Service(executable_path=chromedriver_path)
            logging.info(f'Using provided chromedriver: {chromedriver_path}')
        else:
            service = Service(ChromeDriverManager().install())
        
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_window_size(1420, 1080)
        try:
            driver.set_page_load_timeout(per_song_timeout)
            driver.set_script_timeout(int(CFG.get("SCRIPT_TIMEOUT_SEC", per_song_timeout)))
            try:
                driver.command_executor.set_timeout(int(CFG.get("WEBDRIVER_CMD_TIMEOUT_SEC", 600)))
            except Exception:
                pass
        except Exception:
            pass
        try:
            driver.execute_script("document.body.style.zoom='90%'")
        except Exception:
            pass
        if driver.service and driver.service.process:
            webdriver_processes.append(driver.service.process.pid)
        try:
            launched_user_dirs.append(ud_dir)
        except Exception:
            pass
        return driver
    except Exception as e:
        msg = str(e)
        logging.error(f'WebDriverの初期化に失敗: {msg}')
        if (not _is_retry) and ("DevToolsActivePort" in msg or "session not created" in msg):
            logging.warning('DevToolsActivePort/session not created エラーのため、headless + Chrome再起動で再試行します。')
            try:
                kill_chrome_processes()
            except Exception:
                pass
            return init_driver(
                headless=True,
                per_song_timeout=per_song_timeout,
                for_function1=for_function1,
                profile_name=profile_name,
                chromedriver_path=chromedriver_path,
                _is_retry=True
            )
        try:
            messagebox.showerror('エラー', f'WebDriverの初期化に失敗しました: {msg}')
        except Exception:
            pass
        return None


def enable_speed_blocking(driver):
    try:
        driver.execute_cdp_cmd("Network.enable", {})
        patterns = [
            "*.png", "*.jpg", "*.jpeg", "*.gif", "*.webp", "*.svg", "*.ico",
            "*.mp4", "*.m4v", "*.mov", "*.m3u8", "*.ts", "*.avi", "*.wmv", "*.flv", "*.webm",
            "*.woff", "*.woff2", "*.ttf", "*.otf"
        ]
        driver.execute_cdp_cmd("Network.setBlockedURLs", {"urls": patterns})
        logging.info("[機能2] 重いリソースをブロックしました")
    except Exception as e:
        logging.info(f"[機能2] リソースブロック失敗: {e}")


def safe_quit(driver):
    try:
        if driver:
            driver.quit()
    except Exception:
        pass


# === 【追加】ログイン状態確認機能 ===========================
def check_tiktok_login_status(driver) -> tuple[bool, str]:
    """
    TikTokにログインしているかどうかを確認する。
    
    戻り値: (ログイン済みかどうか, ユーザー名または状態メッセージ)
    """
    try:
        # TikTokのホームページにアクセス
        driver.get("https://www.tiktok.com/")
        time.sleep(3)
        
        # ページ読み込み待機
        for _ in range(10):
            try:
                rs = driver.execute_script("return document.readyState")
                if rs in ("interactive", "complete"):
                    break
            except Exception:
                pass
            time.sleep(0.5)
        
        time.sleep(2)
        
        # ログイン状態を複数の方法で確認
        
        # 方法1: ログインボタンの存在確認（存在すれば未ログイン）
        login_button_selectors = [
            '[data-e2e="top-login-button"]',
            'button[class*="LoginButton"]',
            'a[href*="/login"]',
            '//button[contains(text(), "ログイン")]',
            '//a[contains(text(), "ログイン")]',
        ]
        
        for sel in login_button_selectors:
            try:
                if sel.startswith('//'):
                    els = driver.find_elements(By.XPATH, sel)
                else:
                    els = driver.find_elements(By.CSS_SELECTOR, sel)
                if els:
                    for el in els:
                        if el.is_displayed():
                            logging.info(f'[ログイン確認] ログインボタンが見つかりました（未ログイン状態）')
                            return (False, "未ログイン")
            except Exception:
                pass
        
        # 方法2: プロフィールアイコンやユーザー名の存在確認（存在すればログイン済み）
        profile_selectors = [
            '[data-e2e="profile-icon"]',
            'div[class*="DivAvatarContainer"]',
            'img[class*="ImgAvatar"]',
            '[data-e2e="nav-profile"]',
        ]
        
        for sel in profile_selectors:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                if els:
                    for el in els:
                        if el.is_displayed():
                            logging.info(f'[ログイン確認] プロフィールアイコンが見つかりました（ログイン済み）')
                            return (True, "ログイン済み")
            except Exception:
                pass
        
        # 方法3: Cookieの確認
        try:
            cookies = driver.get_cookies()
            session_cookies = [c for c in cookies if 'sessionid' in c.get('name', '').lower() or 'sid_tt' in c.get('name', '').lower()]
            if session_cookies:
                logging.info(f'[ログイン確認] セッションCookieが見つかりました（ログイン済みの可能性）')
                return (True, "ログイン済み（Cookie確認）")
        except Exception:
            pass
        
        # 判定できない場合
        logging.warning('[ログイン確認] ログイン状態を判定できませんでした')
        return (False, "判定不能")
        
    except Exception as e:
        logging.error(f'[ログイン確認] エラー: {e}')
        return (False, f"エラー: {e}")


def wait_for_manual_login(driver, timeout_sec: int = 120) -> bool:
    """
    手動ログインを待機する。
    ユーザーがブラウザでログインするのを待つ。
    
    戻り値: ログインが完了したかどうか
    """
    logging.info(f'[ログイン待機] TikTokにログインしてください（最大{timeout_sec}秒待機）')
    
    # ログインページに遷移
    try:
        driver.get("https://www.tiktok.com/login")
    except Exception:
        pass
    
    start_time = time.time()
    check_interval = 5  # 5秒ごとにチェック
    
    while time.time() - start_time < timeout_sec:
        if is_stopped():
            return False
        
        time.sleep(check_interval)
        
        is_logged_in, status = check_tiktok_login_status(driver)
        if is_logged_in:
            logging.info(f'[ログイン待機] ログイン確認: {status}')
            return True
        
        remaining = int(timeout_sec - (time.time() - start_time))
        logging.info(f'[ログイン待機] ログイン待機中... 残り{remaining}秒')
    
    logging.warning('[ログイン待機] タイムアウト：ログインが確認できませんでした')
    return False


def ensure_logged_in(driver, headless: bool = False) -> bool:
    """
    ログイン状態を確認し、未ログインなら手動ログインを促す。
    headlessモードの場合は確認のみ。
    
    戻り値: ログイン済みまたはログイン成功したかどうか
    """
    is_logged_in, status = check_tiktok_login_status(driver)
    
    if is_logged_in:
        logging.info(f'[ログイン確認] ログイン済み: {status}')
        return True
    
    logging.warning(f'[ログイン確認] 未ログイン状態: {status}')
    
    if headless:
        logging.error('[ログイン確認] headlessモードでは手動ログインできません。事前にログイン済みのプロファイルを使用してください。')
        return False
    
    # GUIでダイアログを表示
    try:
        result = messagebox.askyesno(
            'TikTok ログイン確認',
            'TikTokにログインしていません。\n\n'
            '開いたブラウザウィンドウでTikTokにログインしてください。\n'
            'ログイン後、「はい」をクリックして続行してください。\n\n'
            '続行しますか？'
        )
        
        if not result:
            logging.info('[ログイン確認] ユーザーがキャンセルしました')
            return False
        
        # ログイン待機
        timeout = int(CFG.get("LOGIN_WAIT_TIMEOUT_SEC", 120))
        return wait_for_manual_login(driver, timeout)
        
    except Exception as e:
        logging.error(f'[ログイン確認] ダイアログ表示エラー: {e}')
        return False


# === エラーページ検知 ======================
ERROR_XPATHS_STRICT = [
    '//p[contains(text(), "不明なエラーが発生しました")]',
    '//p[contains(text(), "ページを表示できません")]',
    '//p[contains(text(), "動画は現在ご利用できません")]',
    '//div[contains(@class, "error") and contains(text(), "エラー")]',
]


def _detect_error_page(driver) -> bool:
    try:
        for xp in ERROR_XPATHS_STRICT:
            els = driver.find_elements(By.XPATH, xp)
            if els:
                for el in els:
                    try:
                        if el.is_displayed():
                            return True
                    except Exception:
                        pass
    except Exception:
        pass
    return False


def _check_video_list_exists(driver, timeout: float = 10.0) -> bool:
    t0 = time.time()
    while time.time() - t0 < timeout:
        try:
            selectors = [
                '[data-e2e="music-item-list"] a[href*="/video/"]',
                '[data-e2e="music-item-list"] a[href*="/photo/"]',
                '[data-e2e="music-item-list"] div[class*="DivItemContainer"]',
                'div[class*="DivVideoFeedV2"] a[href*="/video/"]',
                'a[href*="/@"][href*="/video/"]',
            ]
            for sel in selectors:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                if els and len(els) > 0:
                    logging.info(f'[動画リスト確認] セレクタ "{sel}" で {len(els)} 件発見')
                    return True
        except Exception as e:
            logging.debug(f'動画リスト確認中の例外: {e}')
        time.sleep(0.5)
    return False


def _wait_for_page_load(driver, timeout: float = 10.0) -> bool:
    t0 = time.time()
    while time.time() - t0 < timeout:
        try:
            rs = driver.execute_script("return document.readyState")
            if rs in ("interactive", "complete"):
                body = driver.find_elements(By.TAG_NAME, "body")
                if body:
                    return True
        except Exception:
            pass
        time.sleep(0.2)
    return False


def resolve_final_url(driver, url: str, max_retry: int = 3, sleep_sec: float = 2.0) -> tuple[str, bool]:
    target = url.strip()
    
    for i in range(max_retry):
        try:
            driver.get(target)
        except TimeoutException:
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass
        
        wait_time = sleep_sec + i * 0.5
        time.sleep(wait_time)
        _wait_for_page_load(driver, timeout=5.0)
        
        if _detect_error_page(driver):
            logging.info(f"[正規化] エラーページ検知 → refresh (try {i + 1}/{max_retry})")
            
            extra_wait = int(CFG.get("ERROR_PAGE_EXTRA_WAIT_SEC", 10))
            logging.info(f"[正規化] {extra_wait}秒待機後にリフレッシュします")
            time.sleep(extra_wait)
            
            try:
                driver.refresh()
            except Exception:
                pass
            time.sleep(sleep_sec)
            continue
        
        if _check_video_list_exists(driver, timeout=5.0):
            logging.info('[正規化] 動画リスト確認OK → 正常なページと判断')
            try:
                cur = driver.current_url or target
                return (cur, False)
            except Exception:
                pass
        
        try:
            cur = driver.current_url or target
            return (cur, False)
        except Exception:
            pass
    
    return (target, True)


def wait_dom_interactive(driver, timeout: float = 3.0) -> bool:
    t0 = time.time()
    while time.time() - t0 < timeout:
        try:
            rs = driver.execute_script("return document.readyState")
            if rs in ("interactive", "complete"):
                return True
        except Exception:
            pass
        time.sleep(0.1)
    return False


def wait_rehydration_json(driver, timeout: float = 4.0) -> bool:
    t0 = time.time()
    while time.time() - t0 < timeout:
        try:
            ok = driver.execute_script(
                "var el=document.getElementById('__UNIVERSAL_DATA_FOR_REHYDRATION__');"
                "return !!(el && el.innerText && el.innerText.length>10);"
            )
            if ok:
                return True
        except Exception:
            pass
        time.sleep(0.15)
    return False


# === Excel I/O ==================================
def create_backup_if_exists(excel_filename: str):
    try:
        if os.path.exists(excel_filename):
            backup_dir = os.path.join(exec_dir, 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            current_date = datetime.now().strftime("%Y%m%d")
            base_filename = os.path.basename(excel_filename)
            name, ext = os.path.splitext(base_filename)
            backup_filename = f"{name}_{current_date}{ext}"
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copy2(excel_filename, backup_path)
            cutoff = datetime.now() - timedelta(days=14)
            for fn in os.listdir(backup_dir):
                if not fn.startswith(name) or not fn.endswith(ext):
                    continue
                date_part = fn[len(name) + 1:-len(ext)]
                try:
                    d = datetime.strptime(date_part, "%Y%m%d")
                    if d < cutoff:
                        os.remove(os.path.join(backup_dir, fn))
                except Exception:
                    pass
            logging.info(f'バックアップ作成: {backup_path}')
    except Exception as e:
        logging.warning(f'バックアップ作成中のエラー: {e}')


def create_excel_file(filename: str):
    wb = Workbook()
    sh = wb.active
    sh.title = '楽曲情報'
    sh.append(['楽曲名', '楽曲URL', '日付', '総UGC数'])
    sh.column_dimensions['A'].width = 30
    sh.column_dimensions['B'].width = 20
    sh.column_dimensions['C'].width = 18
    sh.column_dimensions['D'].width = 9

    icon = wb.create_sheet(title='ユーザーアイコン')
    icon.append(['アカウント名', 'アイコンパス'])
    icon.column_dimensions['A'].width = 25
    icon.column_dimensions['B'].width = 40

    wb.save(filename)
    return wb


def update_music_info_sheet(driver, workbook, song_name: str, song_url: str,
                            excel_filename: str, per_song_timeout: int = 300):
    sheet = workbook['楽曲情報']
    logging.info(f'[機能1] 楽曲ページを開きます: {song_name} | {song_url}')

    total_ugc = 0

    try:
        driver.set_page_load_timeout(per_song_timeout)
    except Exception:
        pass

    final_url, _ = resolve_final_url(driver, song_url)

    time.sleep(random.uniform(1, 2))

    try:
        try:
            title_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@data-e2e="music-title"]'))
            )
            fetched_song_name = (title_element.text or '').strip() or song_name
        except TimeoutException:
            fetched_song_name = song_name

        try:
            el = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@data-e2e="music-video-count"]'))
            )
            txt = (el.text or '').replace('動画', '').replace(',', '').strip()
            total_ugc = parse_number(txt) or 0
        except TimeoutException:
            logging.warning('[機能1] 総UGC数の取得に失敗')

        today_date_only = datetime.today().date()

        found = False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == fetched_song_name:
                found = True
                row[1].value = final_url
                row[2].value = today_date_only.strftime('%Y/%m/%d')
                row[3].value = total_ugc
                break

        if not found:
            sheet.append([fetched_song_name, final_url,
                          today_date_only.strftime('%Y/%m/%d'), total_ugc])

    except Exception as e:
        logging.error(f'楽曲情報の更新中にエラー: {e}')


def create_or_clear_date_sheet(workbook, date_str: str | None = None):
    date_sheet_name = date_str or datetime.today().strftime('%Y%m%d')
    if date_sheet_name in workbook.sheetnames:
        sheet = workbook[date_sheet_name]
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
    else:
        sheet = workbook.create_sheet(title=date_sheet_name)
        sheet.append([
            '投稿ID', '投稿日', 'アカウント名', 'ニックネーム',
            'いいね数', 'コメント数', '保存数', 'シェア数',
            '再生回数', 'フォロワー数', '動画リンク(URL)', '更新日'
        ])
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 13
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 26
        sheet.column_dimensions['L'].width = 18
    return sheet


def ensure_date_sheet_exists(workbook, date_str: str):
    if date_str in workbook.sheetnames:
        return workbook[date_str]
    like_dates = [s for s in workbook.sheetnames if re.fullmatch(r'\d{8}', s)]
    if like_dates:
        like_dates.sort()
        return workbook[like_dates[-1]]
    ws = workbook.create_sheet(title=date_str)
    ws.append([
        '投稿ID', '投稿日', 'アカウント名', 'ニックネーム',
        'いいね数', 'コメント数', '保存数', 'シェア数',
        '再生回数', 'フォロワー数', '動画リンク(URL)', '更新日'
    ])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 26
    ws.column_dimensions['L'].width = 18
    return ws


def get_video_urls(driver, max_items=None, timeout=10):
    video_urls: list[str] = []
    seen: set[str] = set()

    try:
        driver.set_script_timeout(int(CFG.get("SCRIPT_TIMEOUT_SEC", 300)))
    except Exception:
        pass

    scroll_attempts = 0
    max_scroll_attempts = 5

    while True:
        if is_stopped():
            logging.info('動画URLの取得が停止されました')
            break

        try:
            prev_height = driver.execute_script("return document.body.scrollHeight")
        except (WebDriverException, ProtocolError, MaxRetryError, ReadTimeoutError, TimeoutException) as e:
            if _is_invalid_session_error(e):
                logging.error(f'ブラウザセッションが無効になったためURL取得を中断します: {e}')
                break
            logging.warning(f'高さ取得に失敗（再試行）: {e}')
            time.sleep(1.0)
            continue
        except Exception as e:
            logging.warning(f'高さ取得で想定外の例外: {e}')
            time.sleep(1.0)
            continue

        for _ in range(10):
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            except (WebDriverException, ProtocolError, MaxRetryError, ReadTimeoutError, TimeoutException) as e:
                if _is_invalid_session_error(e):
                    logging.error(f'ブラウザセッションが無効になったためURL取得を中断します: {e}')
                    return video_urls
            except Exception as e:
                logging.warning(f'スクロールで想定外の例外（継続）: {e}')
                break
            time.sleep(random.uniform(1.0, 2.0))

        try:
            new_height = driver.execute_script("return document.body.scrollHeight")
        except (WebDriverException, ProtocolError, MaxRetryError, ReadTimeoutError, TimeoutException) as e:
            if _is_invalid_session_error(e):
                logging.error(f'ブラウザセッションが無効になったためURL取得を中断します: {e}')
                break
            logging.warning(f'高さ再取得に失敗: {e}')
            new_height = prev_height
        except Exception as e:
            logging.warning(f'高さ再取得で想定外の例外: {e}')
            new_height = prev_height

        try:
            all_urls = driver.execute_script("""
                const anchors = Array.from(
                  document.querySelectorAll(
                    '[data-e2e="music-item-list"] a[href*="/video/"], ' +
                    '[data-e2e="music-item-list"] a[href*="/photo/"], ' +
                    'a[href*="/video/"], a[href*="/photo/"]'
                  )
                );
                const urls = anchors.map(a => a.href);
                return Array.from(new Set(urls));
            """)
        except (WebDriverException, ProtocolError, MaxRetryError, ReadTimeoutError, TimeoutException) as e:
            if _is_invalid_session_error(e):
                logging.error(f'ブラウザセッションが無効になったためURL取得を中断します: {e}')
                break
            logging.warning(f'URL抽出スクリプト実行エラー（継続）: {e}')
            continue
        except Exception as e:
            logging.warning(f'URL抽出で想定外の例外（継続）: {e}')
            continue

        new_urls = [u for u in all_urls if u and u not in seen]

        if new_urls:
            video_urls.extend(new_urls)
            for u in new_urls:
                seen.add(u)
            logging.info(f'現在の取得件数: {len(video_urls)}')
            scroll_attempts = 0
        else:
            if new_height == prev_height:
                scroll_attempts += 1
                logging.info(f'新しい投稿が見つからない試行: {scroll_attempts}/{max_scroll_attempts}')
                if scroll_attempts >= max_scroll_attempts:
                    logging.info('これ以上新規投稿が読み込まれないため終了')
                    break
            else:
                scroll_attempts = 0

        if max_items and len(video_urls) >= max_items:
            break

    logging.info(f'最終取得動画URL件数: {len(video_urls)}')
    return video_urls


def write_video_links(sheet, video_urls):
    for url in video_urls:
        sheet.append([None] * 10 + [url] + [None])


def write_update_dates(sheet):
    now = datetime.today().strftime('%Y/%m/%d %H:%M')
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            if not cell.value:
                cell.value = now


# === 投稿日/統計値 取得 =============================
def _safe_attr(el, attr, default=''):
    try:
        return el.get_attribute(attr)
    except Exception:
        return default


def _video_id_from_url(url: str) -> str:
    m = re.search(r'/video/(\d+)', url)
    return m.group(1) if m else (url.rstrip('/').split('/')[-1].split('?')[0])


def extract_date_from_span_pattern1(driver):
    try:
        span = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.XPATH, '//span[@data-e2e="browser-nickname"]'))
        )
        spans = span.find_elements(By.XPATH, './span')
        if len(spans) >= 3:
            return spans[2].text
    except Exception:
        return None
    return None


def extract_date_from_span_pattern2(driver):
    try:
        elems = WebDriverWait(driver, 3).until(
            EC.presence_of_all_elements_located((By.XPATH, '//span[contains(@class, "SpanOtherInfos")]'))
        )
        for el in elems:
            t = (el.text or '')
            if '·' in t:
                parts = t.split('·')
                if len(parts) >= 2:
                    return parts[1].strip()
    except Exception:
        return None
    return None


def extract_date_from_json_data(driver):
    try:
        js = driver.execute_script(
            "var el=document.getElementById('__UNIVERSAL_DATA_FOR_REHYDRATION__');"
            "return el?el.innerText:'';"
        )
        if not js:
            return None
        data = json.loads(js)
        create_time = (
            data.get("__DEFAULT_SCOPE__", {})
            .get("webapp.video-detail", {})
            .get("itemInfo", {})
            .get("itemStruct", {})
            .get("createTime")
        )
        if create_time:
            dt = datetime.fromtimestamp(int(create_time))
            return f"{dt.year}/{dt.month:02d}/{dt.day:02d}"
    except Exception:
        return None
    return None


def extract_date(driver):
    for fn in (extract_date_from_span_pattern1, extract_date_from_span_pattern2, extract_date_from_json_data):
        try:
            r = fn(driver)
            if r:
                return r
        except Exception:
            pass
    return ''


def parse_date_posted(date_posted_str, update_datetime):
    s0 = (date_posted_str or '').strip()
    if not s0:
        return ''
    s = re.sub(r'\s+', '', s0)
    s = s.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
    if s in ('今日', 'きょう', 'today'):
        return update_datetime.strftime('%Y/%m/%d')
    if s in ('昨日', 'きのう', 'yesterday'):
        return (update_datetime - timedelta(days=1)).strftime('%Y/%m/%d')
    m = re.match(r'^(\d+)分前$', s)
    if m:
        return update_datetime.strftime('%Y/%m/%d')
    m = re.match(r'^(\d+)時間前$', s)
    if m:
        return (update_datetime - timedelta(hours=int(m.group(1)))).strftime('%Y/%m/%d')
    m = re.match(r'^(\d+)日前$', s)
    if m:
        return (update_datetime - timedelta(days=int(m.group(1)))).strftime('%Y/%m/%d')
    m = re.match(r'^(\d+)週間前$', s)
    if m:
        return (update_datetime - timedelta(weeks=int(m.group(1)))).strftime('%Y/%m/%d')
    s_abs = (s.replace('－', '-').replace('—', '-').replace('–', '-')
             .replace('／', '/').replace('-', '/').replace('.', '/'))
    s_abs = s_abs.replace('年', '/').replace('月', '/').replace('日', '')
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', s_abs)
    if m:
        y, mn, dy = map(int, m.groups())
        return f'{y:04d}/{mn:02d}/{dy:02d}'
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', s_abs)
    if m:
        mn, dy = map(int, m.groups())
        y = update_datetime.year
        if (mn, dy) > (update_datetime.month, update_datetime.day):
            y -= 1
        return f'{y:04d}/{mn:02d}/{dy:02d}'
    return date_posted_str


def extract_video_stats_from_json(driver):
    try:
        script_content = driver.execute_script(
            "var el=document.getElementById('__UNIVERSAL_DATA_FOR_REHYDRATION__');"
            "return el?el.innerText:'';"
        )
        data = json.loads(script_content) if script_content else None
        if not data:
            raise ValueError('rehydration json not found')

        item = (data["__DEFAULT_SCOPE__"]["webapp.video-detail"]["itemInfo"]["itemStruct"])
        stats = item["stats"]
        author_stats = item.get("authorStats", {})
        author = item.get("author", {})

        return {
            'いいね数': int(stats.get("diggCount") or 0),
            'コメント数': int(stats.get("commentCount") or 0),
            'シェア数': int(stats.get("shareCount") or 0),
            '再生回数': int(stats.get("playCount") or 0),
            '保存数': int(stats.get("collectCount") or 0),
            'フォロワー数': int(author_stats.get("followerCount") or 0),
            'アカウント名': author.get("uniqueId") or '',
            'ニックネーム': author.get("nickname") or ''
        }
    except Exception as e:
        logging.info(f'JSON直読み失敗（フォールバックへ）: {e}')
        return {
            'いいね数': '', 'コメント数': '', 'シェア数': '', '再生回数': '',
            '保存数': '', 'フォロワー数': '', 'アカウント名': '', 'ニックネーム': ''
        }


def extract_text_with_retry(driver, by, value, description):
    try:
        element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((by, value)))
        return element.text or ''
    except Exception:
        return ''


def extract_and_parse_number(driver, xpath, description):
    text = extract_text_with_retry(driver, By.XPATH, xpath, description)
    if text:
        v = parse_number(text)
        return '' if v is None else v
    return ''


prefer_account_name_alt = True


def extract_avatar_url(driver, account_name):
    global prefer_account_name_alt
    try:
        if prefer_account_name_alt:
            img = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f'//img[@alt="{account_name}"]'))
            )
        else:
            img = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//a[@data-e2e="browse-user-avatar"]//img[@alt=""]')
                )
            )
        return _safe_attr(img, 'src', '')
    except TimeoutException:
        if prefer_account_name_alt:
            prefer_account_name_alt = False
            return extract_avatar_url(driver, account_name)
        return ''
    except Exception:
        return ''


def get_follower_count(driver, account_name, original_window, follower_window_handle):
    follower_count = ''
    nickname = ''
    try:
        account_url = f'https://www.tiktok.com/@{account_name}'
        driver.switch_to.window(follower_window_handle)
        driver.get(account_url)
        wait_dom_interactive(driver, 2.5)
        nickname = extract_text_with_retry(
            driver, By.XPATH, '//h2[@data-e2e="user-subtitle"]', 'ニックネーム'
        )
        el = WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.XPATH, '//strong[@data-e2e="followers-count"]'))
        )
        follower_count = parse_number(el.text)
        driver.switch_to.window(original_window)
    except Exception as e:
        logging.info(f'プロフィールからのフォロワー/ニックネーム取得失敗: {e}')
    return follower_count or '', nickname or ''


def write_video_data_to_row(sheet, row, data):
    ordered_keys = [
        '投稿ID', '投稿日', 'アカウント名', 'ニックネーム', 'いいね数', 'コメント数',
        '保存数', 'シェア数', '再生回数', 'フォロワー数', '動画リンク(URL)', '更新日'
    ]
    for i, key in enumerate(ordered_keys):
        row[i].value = data.get(key, '')


def extract_video_data(driver, original_window, follower_window_handle,
                       follower_cache: dict[str, tuple] | None = None):
    data: dict[str, object] = {}
    try:
        current_url = driver.current_url
        post_id = _video_id_from_url(current_url)
        data['投稿ID'] = post_id

        raw_date = (extract_date(driver) or '').strip()
        update_dt = datetime.now()
        data['投稿日'] = parse_date_posted(raw_date, update_dt)
        data['更新日'] = datetime.today().strftime('%Y/%m/%d %H:%M')

        if '/video/' in current_url:
            json_stats = extract_video_stats_from_json(driver)
            data.update(json_stats)

            if not data.get('アカウント名'):
                m = re.search(r'/@([^/]+)/', current_url)
                if m:
                    data['アカウント名'] = m.group(1)

            account = (data.get('アカウント名') or '').strip()
            need_profile = (not data.get('ニックネーム') or not data.get('フォロワー数'))
            if need_profile and account:
                if follower_cache is not None and account in follower_cache:
                    fol, nick = follower_cache[account]
                else:
                    fol, nick = get_follower_count(driver, account, original_window, follower_window_handle)
                    if follower_cache is not None:
                        follower_cache[account] = (fol, nick)
                if nick and not data.get('ニックネーム'):
                    data['ニックネーム'] = nick
                if fol and not data.get('フォロワー数'):
                    data['フォロワー数'] = fol
        else:
            m = re.search(r'/@([^/]+)/', current_url)
            data['アカウント名'] = m.group(1) if m else ''
            data['いいね数'] = extract_and_parse_number(driver, '//strong[@data-e2e="like-count"]', 'いいね数')
            data['コメント数'] = extract_and_parse_number(driver, '//strong[@data-e2e="comment-count"]', 'コメント数')
            data['保存数'] = extract_and_parse_number(driver, '//strong[@data-e2e="undefined-count"]', '保存数')
            share = extract_and_parse_number(driver, '//strong[@data-e2e="share-count"]', 'シェア数')
            data['シェア数'] = 0 if share == '' else share
            data['再生回数'] = ''

            account = data.get('アカウント名') or ''
            if account:
                if follower_cache is not None and account in follower_cache:
                    fol, nick = follower_cache[account]
                else:
                    fol, nick = get_follower_count(driver, account, original_window, follower_window_handle)
                    if follower_cache is not None:
                        follower_cache[account] = (fol, nick)
                data['フォロワー数'] = fol or ''
                data['ニックネーム'] = nick or ''

        data['動画リンク(URL)'] = current_url
        account_for_avatar = data.get('アカウント名') or ''
        data['アバターURL'] = extract_avatar_url(driver, account_for_avatar) if account_for_avatar else ''

    except Exception as e:
        logging.error(f'動画データ抽出エラー: {e}', exc_info=True)

    for k in ['投稿ID', '投稿日', 'アカウント名', 'ニックネーム', 'いいね数', 'コメント数',
              '保存数', 'シェア数', '再生回数', 'フォロワー数', '動画リンク(URL)', '更新日', 'アバターURL']:
        data.setdefault(k, '')

    return data


# === 【修正v3】機能1：URL収集 ==============================================
def function1(save_path, max_items, song_urls, headless=False,
              per_song_timeout: int = 300, skip_on_timeout: bool = False):
    logging.info('#機能1 開始 (timeout=%ss, skip_on_timeout=%s)', per_song_timeout, skip_on_timeout)

    consecutive_errors = 0
    max_consecutive_errors = int(CFG.get("MAX_CONSECUTIVE_ERRORS", 3))
    song_interval_min = int(CFG.get("SONG_INTERVAL_MIN_SEC", 8))
    song_interval_max = int(CFG.get("SONG_INTERVAL_MAX_SEC", 15))
    long_wait_sec = int(CFG.get("LONG_WAIT_AFTER_ERRORS_SEC", 60))

    # 【追加】最初にログイン確認用のドライバを起動
    check_login = bool(CFG.get("CHECK_LOGIN_BEFORE_START", True))
    if check_login:
        logging.info('[機能1] ログイン状態を確認します...')
        login_driver = init_driver(headless=headless, per_song_timeout=per_song_timeout, for_function1=True)
        if login_driver:
            try:
                if not ensure_logged_in(login_driver, headless=headless):
                    logging.error('[機能1] ログインが確認できなかったため、処理を中断します')
                    safe_quit(login_driver)
                    return
                logging.info('[機能1] ログイン確認完了')
            finally:
                safe_quit(login_driver)
            # ログイン確認後、少し待機
            time.sleep(2)

    for idx, (song_name, song_url) in enumerate(song_urls):
        if is_stopped():
            logging.info('#機能1 停止フラグ検知 → 中断')
            break

        if idx > 0:
            wait_sec = random.uniform(song_interval_min, song_interval_max)
            logging.info(f'[機能1] Rate limiting対策: 次の曲まで {wait_sec:.1f}秒待機します')
            time.sleep(wait_sec)

        if consecutive_errors >= max_consecutive_errors:
            logging.warning(f'[機能1] {consecutive_errors}回連続でエラー → {long_wait_sec}秒の長い待機を実行')
            time.sleep(long_wait_sec)
            consecutive_errors = 0

        driver = init_driver(headless=headless, per_song_timeout=per_song_timeout, for_function1=True)
        if not driver:
            logging.error('WebDriver初期化に失敗（機能1）')
            break

        try:
            sanitized = sanitize_filename(song_name)
            excel_filename = os.path.join(save_path, f'{sanitized}.xlsx')

            create_backup_if_exists(excel_filename)
            if not os.path.exists(excel_filename):
                wb = create_excel_file(excel_filename)
            else:
                wb = load_workbook(excel_filename)
                if 'ユーザーアイコン' not in wb.sheetnames:
                    icon = wb.create_sheet(title='ユーザーアイコン')
                    icon.append(['アカウント名', 'アイコンパス'])
                    icon.column_dimensions['A'].width = 25
                    icon.column_dimensions['B'].width = 40

            final_url, had_fatal_error = resolve_final_url(driver, song_url)
            
            update_music_info_sheet(driver, wb, song_name, final_url, excel_filename, per_song_timeout)

            today_key = datetime.today().strftime('%Y%m%d')
            date_sheet = create_or_clear_date_sheet(wb, today_key)

            try:
                driver.get(final_url)
            except TimeoutException:
                logging.warning(f'[機能1] ページ遷移タイムアウト: {song_name} → 続行')
                try:
                    driver.execute_script("window.stop();")
                except Exception:
                    pass

            time.sleep(random.uniform(3, 5))
            _wait_for_page_load(driver, timeout=10.0)

            urls = []
            skip_this_song = False

            logging.info(f'[機能1] 動画リストの確認中: {song_name}')
            has_video_list = _check_video_list_exists(driver, timeout=15.0)
            
            if not has_video_list:
                if _detect_error_page(driver):
                    logging.warning(f'[機能1] エラーページ検知: {song_name} → リフレッシュ試行')
                    
                    extra_wait = int(CFG.get("ERROR_PAGE_EXTRA_WAIT_SEC", 10))
                    time.sleep(extra_wait)
                    
                    try:
                        driver.refresh()
                    except Exception:
                        pass
                    time.sleep(5.0)
                    _wait_for_page_load(driver, timeout=10.0)
                    
                    has_video_list = _check_video_list_exists(driver, timeout=10.0)
                    if not has_video_list:
                        if _detect_error_page(driver):
                            logging.warning(f'[機能1] 楽曲ページがエラー表示のためURL収集をスキップします: {song_name}')
                            skip_this_song = True
                            consecutive_errors += 1
                        else:
                            logging.warning(f'[機能1] 動画リストが見つかりませんが処理を続行します: {song_name}')
                else:
                    logging.warning(f'[機能1] 動画リストが見つかりません（エラーページではない）: {song_name}')
            
            if has_video_list:
                logging.info(f'[機能1] 動画リスト確認OK: {song_name}')
            
            if not skip_this_song:
                urls = get_video_urls(driver, max_items=max_items, timeout=10)
                
                if urls:
                    consecutive_errors = 0
                    logging.info(f'[機能1] URL取得成功: {song_name} ({len(urls)}件)')
                else:
                    if has_video_list:
                        consecutive_errors += 1
                        logging.warning(f'[機能1] 動画リストはあるがURL取得結果が0件: {song_name}')
                    else:
                        consecutive_errors += 1
                        logging.warning(f'[機能1] URL取得結果が0件: {song_name}')

            if urls:
                write_video_links(date_sheet, urls)
            write_update_dates(date_sheet)

            while True:
                try:
                    wb.save(excel_filename)
                    break
                except PermissionError:
                    try:
                        messagebox.showwarning('警告', 'ファイルが開かれているため保存できません。\n閉じて OK を押してください。')
                    except Exception:
                        pass
                except Exception as e:
                    logging.error(f'保存中エラー（機能1）: {e}')
                    try:
                        messagebox.showerror('エラー', f'保存中にエラーが発生しました: {e}')
                    except Exception:
                        pass
                    break

            logging.info(f'[機能1] {song_name}: 取得URL件数={len(urls)}')

        except TimeoutException as te:
            logging.warning(f'[機能1] タイムアウト: {song_name} | {te}')
            consecutive_errors += 1
            if not skip_on_timeout:
                raise
        except Exception as e:
            logging.error(f'[機能1] 例外: {song_name} | {e}', exc_info=True)
            consecutive_errors += 1
            if not skip_on_timeout:
                raise
        finally:
            safe_quit(driver)

    logging.info('#機能1 終了')


# === 機能2 =====================================
def _function2_core(save_path, song_urls, headless=False,
                    per_song_timeout: int = 300, skip_on_timeout: bool = False,
                    profile_name: str | None = None, work_date_str: str | None = None,
                    chromedriver_path: str | None = None):
    ensure_dir(images_dir)
    driver = init_driver(headless=headless, per_song_timeout=per_song_timeout, profile_name=profile_name, chromedriver_path=chromedriver_path)
    if not driver:
        logging.error('WebDriver初期化に失敗（機能2）')
        raise DriverInitFatalError('WebDriver初期化に失敗（機能2）')

    enable_speed_blocking(driver)

    try:
        try:
            original_window = driver.current_window_handle
            driver.execute_script("window.open('');")
            follower_window_handle = driver.window_handles[-1]
        except Exception:
            follower_window_handle = driver.current_window_handle

        operations_count = 0
        INTERMEDIATE_SAVE_EVERY = 50
        follower_cache: dict[str, tuple] = {}
        session = requests.Session()
        fixed_date_str = work_date_str or datetime.now().strftime('%Y%m%d')

        for song_name, _ in song_urls:
            if is_stopped():
                logging.info('#機能2 停止フラグ検知 → 中断')
                break

            sanitized = sanitize_filename(song_name)
            excel_filename = os.path.join(save_path, f'{sanitized}.xlsx')
            if not os.path.exists(excel_filename):
                logging.error(f'ファイルが存在しません: {excel_filename}')
                continue

            wb = load_workbook(excel_filename)
            sheet = ensure_date_sheet_exists(wb, fixed_date_str)

            if 'ユーザーアイコン' not in wb.sheetnames:
                icon = wb.create_sheet(title='ユーザーアイコン')
                icon.append(['アカウント名', 'アイコンパス'])
                icon.column_dimensions['A'].width = 25
                icon.column_dimensions['B'].width = 40
            icon_sheet = wb['ユーザーアイコン']

            account_icons = {}
            for r in icon_sheet.iter_rows(min_row=2):
                if r[0].value and r[1].value:
                    account_icons[str(r[0].value)] = str(r[1].value)

            total_urls_count = sum(1 for r in sheet.iter_rows(min_row=2, min_col=11, max_col=11) if r[0].value)
            song_written = 0

            for row in sheet.iter_rows(min_row=2, max_col=12):
                if is_stopped():
                    logging.info('#機能2 行ループで停止フラグ検知 → 残り行の処理を中断')
                    break

                if all(cell.value not in (None, '') for cell in row[:10]) and row[10].value and row[11].value:
                    continue

                link = row[10].value
                if not link:
                    continue

                driver.switch_to.window(original_window)
                try:
                    driver.get(link)
                except TimeoutException:
                    try:
                        driver.execute_script("window.stop();")
                    except Exception:
                        pass

                wait_dom_interactive(driver, 2.0)
                wait_rehydration_json(driver, 2.5)

                if _detect_error_page(driver):
                    try:
                        driver.refresh()
                    except Exception:
                        pass
                    wait_dom_interactive(driver, 1.5)
                    wait_rehydration_json(driver, 1.5)
                    if _detect_error_page(driver):
                        logging.info('エラーページが続くためスキップ')
                        continue

                data = extract_video_data(driver, original_window, follower_window_handle, follower_cache)
                write_video_data_to_row(sheet, row, data)

                account_name = data.get('アカウント名')
                avatar_url = data.get('アバターURL')
                if account_name and avatar_url and account_name not in account_icons:
                    try:
                        icon_filename = f"{sanitize_filename(account_name)}.jpg"
                        icon_path = os.path.join(images_dir, icon_filename)
                        resp = session.get(avatar_url, stream=True, timeout=5)
                        if resp.status_code == 200:
                            with open(icon_path, 'wb') as f:
                                for chunk in resp.iter_content(1024):
                                    f.write(chunk)
                            account_icons[account_name] = f"images/{icon_filename}"
                            logging.info(f'アイコン保存: {account_name} -> images/{icon_filename}')
                    except Exception as e:
                        logging.info(f'アイコン保存失敗: {e}')

                operations_count += 1
                song_written += 1

                if operations_count % INTERMEDIATE_SAVE_EVERY == 0:
                    while True:
                        try:
                            wb.save(excel_filename)
                            break
                        except PermissionError:
                            try:
                                messagebox.showwarning('警告', 'ファイルが開かれているため保存できません。\n閉じて OK を押してください。')
                            except Exception:
                                pass
                        except Exception as e:
                            logging.error(f'保存中エラー（機能2途中）: {e}')
                            break

            if icon_sheet.max_row > 1:
                icon_sheet.delete_rows(2, icon_sheet.max_row - 1)
            for acc, p in account_icons.items():
                icon_sheet.append([acc, p])

            logging.info(f'{song_name} - 最終取得件数: {song_written}, 総URL件数: {total_urls_count}')

            while True:
                try:
                    wb.save(excel_filename)
                    break
                except PermissionError:
                    try:
                        messagebox.showwarning('警告', 'ファイルが開かれているため保存できません。\n閉じて OK を押してください。')
                    except Exception:
                        pass
                except Exception as e:
                    logging.error(f'保存中エラー（機能2）: {e}')
                    try:
                        messagebox.showerror('エラー', f'保存中にエラーが発生しました: {e}')
                    except Exception:
                        pass
                    break
    finally:
        safe_quit(driver)


def function2(save_path, song_urls, headless=False,
              per_song_timeout: int = 300, skip_on_timeout: bool = False,
              profile_name: str | None = None, work_date_str: str | None = None,
              chromedriver_path: str | None = None):
    logging.info('#機能2 開始 (timeout=%ss, skip_on_timeout=%s)', per_song_timeout, skip_on_timeout)
    try:
        _function2_core(save_path, song_urls, headless=headless, per_song_timeout=per_song_timeout,
                        skip_on_timeout=skip_on_timeout, profile_name=profile_name, work_date_str=work_date_str,
                        chromedriver_path=chromedriver_path)
    finally:
        logging.info('#機能2 終了')


def _function2_worker(payload: dict):
    global external_stop
    external_stop = payload.get("shared_stop")
    try:
        time.sleep(random.uniform(0.2, 0.8))
    except Exception:
        pass

    save_path = payload["save_path"]
    song_name, song_url = payload["song"]
    headless = payload["headless"]
    per_song_timeout = payload["per_song_timeout"]
    skip_on_timeout = payload["skip_on_timeout"]
    profile_name = payload["profile_name"]
    work_date_str = payload.get("work_date_str")
    chromedriver_path = payload.get("chromedriver_path")

    fatal = False
    fatal_reason = ''

    try:
        function2(save_path, [(song_name, song_url)], headless=headless, per_song_timeout=per_song_timeout,
                  skip_on_timeout=skip_on_timeout, profile_name=profile_name, work_date_str=work_date_str,
                  chromedriver_path=chromedriver_path)
    except DriverInitFatalError as e:
        logging.error(f'[並列ワーカー] Chrome起動致命的エラー: {song_name} | {e}')
        fatal = True
        fatal_reason = 'driver_init_failed'
        try:
            if external_stop is not None:
                external_stop.set()
        except Exception:
            pass
    except Exception as e:
        if _is_invalid_session_error(e):
            logging.error(f'[並列ワーカー] セッションエラー(この曲のみ失敗): {song_name} | {e}')
            fatal = False
            fatal_reason = 'invalid_session'
        else:
            logging.error(f'[並列ワーカー] 例外: {song_name} | {e}')

    sanitized = sanitize_filename(song_name)
    excel_filename = os.path.join(save_path, f'{sanitized}.xlsx')
    total_urls = 0
    filled = 0
    date_key = work_date_str or datetime.now().strftime('%Y%m%d')

    try:
        if os.path.exists(excel_filename):
            wb = load_workbook(excel_filename)
            target_sheet = None
            if date_key in wb.sheetnames:
                target_sheet = wb[date_key]
            else:
                like_dates = [s for s in wb.sheetnames if re.fullmatch(r'\d{8}', s)]
                if like_dates:
                    like_dates.sort()
                    target_sheet = wb[like_dates[-1]]
            if target_sheet:
                for row in target_sheet.iter_rows(min_row=2, max_col=12):
                    link = row[10].value
                    if link:
                        total_urls += 1
                        if all(cell.value not in (None, '') for cell in row[:10]) and row[11].value:
                            filled += 1
    except Exception as e:
        logging.info(f'[並列ワーカー] 件数集計で例外: {e}')

    elapsed = 0
    try:
        elapsed = int(payload.get("_tstart") and (time.time() - payload["_tstart"]) or 0)
    except Exception:
        pass

    return {
        "song_name": song_name, "total": total_urls, "filled": filled,
        "elapsed_sec": elapsed, "profile": profile_name, "fatal": fatal, "fatal_reason": fatal_reason
    }


def function2_orchestrator(save_path, song_urls, headless=False,
                           per_song_timeout: int = 300, skip_on_timeout: bool = False):
    logging.info('#機能2 開始 (並列オーケストレータ, timeout=%ss, skip_on_timeout=%s)', per_song_timeout, skip_on_timeout)

    try:
        try:
            chromedriver_path = ChromeDriverManager().install()
            logging.info(f'chromedriverをインストール/確認しました: {chromedriver_path}')
        except Exception as e:
            logging.error(f'chromedriverの事前インストールに失敗: {e}')
            chromedriver_path = None

        unique = []
        seen = set()
        for name, url in song_urls:
            if name not in seen:
                unique.append((name, url))
                seen.add(name)
        if not unique:
            logging.warning('[並列] 対象曲が0件です')
            return

        cfg = CFG
        max_workers = int(cfg.get("PARALLEL_WORKERS", 2))
        profile_pool = list(cfg.get("CHROME_PROFILE_POOL", ["Default"]))
        if not profile_pool:
            profile_pool = ["Default"]
        max_workers = max(1, min(max_workers, len(profile_pool), len(unique)))

        logging.info(f'[並列] 開始: 対象{len(unique)}曲 同時{max_workers}ワーカー')

        failed_songs_for_retry: list[str] = []
        fatal_songs: list[str] = []
        fatal_stop = False

        from multiprocessing import Manager
        mgr = Manager()
        shared_stop = mgr.Event()
        global external_stop
        external_stop = shared_stop

        work_date_str = datetime.now().strftime('%Y%m%d')
        prof_deque = deque(profile_pool[:max_workers])
        tasks_iter = iter(unique)
        futures = {}
        results = []
        from concurrent.futures import CancelledError

        with ProcessPoolExecutor(max_workers=max_workers) as ex:
            while len(futures) < max_workers:
                if is_stopped():
                    break
                try:
                    song = next(tasks_iter)
                except StopIteration:
                    break
                if not prof_deque:
                    break
                prof = prof_deque.popleft()
                payload = {
                    "save_path": save_path, "song": song, "headless": headless,
                    "per_song_timeout": per_song_timeout, "skip_on_timeout": skip_on_timeout,
                    "profile_name": prof, "shared_stop": shared_stop, "work_date_str": work_date_str,
                    "_tstart": time.time(), "chromedriver_path": chromedriver_path
                }
                fut = ex.submit(_function2_worker, payload)
                futures[fut] = prof
                logging.info(f'[並列] START: {song[0]} @ {prof}')

            while futures:
                for fut in as_completed(list(futures.keys())):
                    prof = futures.pop(fut)
                    try:
                        res = fut.result()
                        if res:
                            results.append(res)
                            song_name = res.get("song_name")
                            total = res.get("total", 0)
                            filled = res.get("filled", 0)
                            fatal = bool(res.get("fatal"))
                            fatal_reason = (res.get("fatal_reason") or '')

                            if fatal:
                                fatal_stop = True
                                fatal_songs.append(song_name)
                                logging.error(f'[並列] FATAL: {song_name} Chrome/セッションエラー({fatal_reason})')
                                try:
                                    shared_stop.set()
                                except Exception:
                                    pass
                            else:
                                logging.info(f'[並列] DONE: {song_name} 件数: {filled}/{total} 所要: {res.get("elapsed_sec", 0)}s @ {prof}')
                                if total > 0 and filled == 0:
                                    failed_songs_for_retry.append(song_name)
                    except CancelledError:
                        logging.info(f'[並列] 取消 @ {prof}')
                    except Exception as e:
                        logging.error(f'[並列] ワーカー例外: {e}')

                    prof_deque.append(prof)

                    if is_stopped() or fatal_stop:
                        continue
                    try:
                        song = next(tasks_iter)
                    except StopIteration:
                        continue
                    if not prof_deque:
                        continue
                    prof2 = prof_deque.popleft()
                    payload2 = {
                        "save_path": save_path, "song": song, "headless": headless,
                        "per_song_timeout": per_song_timeout, "skip_on_timeout": skip_on_timeout,
                        "profile_name": prof2, "shared_stop": shared_stop, "work_date_str": work_date_str,
                        "_tstart": time.time(), "chromedriver_path": chromedriver_path
                    }
                    fut2 = ex.submit(_function2_worker, payload2)
                    futures[fut2] = prof2
                    logging.info(f'[並列] START: {song[0]} @ {prof2}')

        if results:
            try:
                total = sum(r.get("total", 0) for r in results)
                filled = sum(r.get("filled", 0) for r in results)
                logging.info(f'[並列] 全体要約: 埋まり {filled}/{total} （{len(results)}曲）')
            except Exception:
                pass
            if failed_songs_for_retry:
                logging.warning(f'[並列] 要リトライ曲: {", ".join(failed_songs_for_retry)}')
            if fatal_songs:
                logging.error(f'[並列] 中断曲: {", ".join(fatal_songs)}')

        logging.info('[並列] 終了')
    finally:
        logging.info('#機能2 終了')


# === GUI ========================================================
def create_gui():
    stop_flag.clear()

    def _guard(fn, *args, **kwargs):
        try:
            fn(*args, **kwargs)
        except Exception as e:
            logging.error(f'実行中にエラー: {e}', exc_info=True)
            try:
                messagebox.showerror('エラー', f'実行中にエラーが発生しました: {e}')
            except Exception:
                pass

    def run_function1():
        save_path, max_items, song_urls = read_initial_settings()
        if save_path and song_urls:
            timeout_cfg = int(CFG.get("PER_SONG_TIME_BUDGET_SEC", 300))
            threading.Thread(target=lambda: _guard(function1, save_path, max_items, song_urls, False, timeout_cfg, True), daemon=True).start()

    def run_function2():
        save_path, _, song_urls = read_initial_settings()
        if save_path and song_urls:
            timeout_cfg = int(CFG.get("PER_SONG_TIME_BUDGET_SEC", 300))

            def runner():
                try:
                    if bool(CFG.get("PARALLEL_ENABLED", False)):
                        function2_orchestrator(save_path, song_urls, headless=False, per_song_timeout=timeout_cfg, skip_on_timeout=True)
                    else:
                        function2(save_path, song_urls, headless=False, per_song_timeout=timeout_cfg, skip_on_timeout=True)
                except Exception as e:
                    logging.error(f'実行中にエラー: {e}', exc_info=True)
                    try:
                        messagebox.showerror('エラー', f'実行中にエラーが発生しました: {e}')
                    except Exception:
                        pass
                finally:
                    kill_chrome_processes()

            threading.Thread(target=runner, daemon=True).start()

    def run_function1_and_2():
        save_path, max_items, song_urls = read_initial_settings()
        if save_path and song_urls:
            timeout_cfg = int(CFG.get("PER_SONG_TIME_BUDGET_SEC", 300))

            def runner():
                try:
                    logging.info('#機能1+2 開始')
                    function1(save_path, max_items, song_urls, headless=False, per_song_timeout=timeout_cfg, skip_on_timeout=True)

                    if is_stopped():
                        logging.info('#機能1+2 停止フラグ検知 → 機能2はスキップします')
                        return

                    if bool(CFG.get("PARALLEL_ENABLED", False)):
                        function2_orchestrator(save_path, song_urls, headless=False, per_song_timeout=timeout_cfg, skip_on_timeout=True)
                    else:
                        function2(save_path, song_urls, headless=False, per_song_timeout=timeout_cfg, skip_on_timeout=True)
                    logging.info('#機能1+2 終了')
                except Exception as e:
                    logging.error(f'実行中にエラー: {e}', exc_info=True)
                    try:
                        messagebox.showerror('エラー', f'実行中にエラーが発生しました: {e}')
                    except Exception:
                        pass
                finally:
                    kill_chrome_processes()

            threading.Thread(target=runner, daemon=True).start()

    def run_stop():
        signal_global_stop()
        kill_chrome_processes()
        force_kill_all_processes()
        try:
            messagebox.showinfo('停止', '全ての処理が停止されました。')
        except Exception:
            pass

    root = tk.Tk()
    root.title('TikTok 集計＆明細ツール v3')
    root.geometry('900x700')

    text_area = scrolledtext.ScrolledText(root, width=100, height=35, state='disabled')
    text_area.grid(row=0, column=0, columnspan=6, padx=10, pady=10)

    log_display_var = tk.BooleanVar(value=True)
    tk.Checkbutton(root, text='ログ表示', variable=log_display_var).grid(row=1, column=0, sticky='w')

    tk.Button(root, text='#機能1（URL収集）', command=run_function1).grid(row=1, column=1, padx=5)
    tk.Button(root, text='#機能2（明細取得）', command=run_function2).grid(row=1, column=2, padx=5)
    tk.Button(root, text='停止', command=run_stop).grid(row=1, column=3, padx=5)
    tk.Button(root, text='機能1 + 機能2', command=run_function1_and_2).grid(row=1, column=4, padx=5)

    text_handler = TextHandler(text_area, log_display_var)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    text_handler.setFormatter(formatter)
    logging.getLogger().addHandler(text_handler)

    root.mainloop()


def slice_song_range(song_urls, offset: int | None, limit: int | None):
    if not song_urls:
        return []
    o = max(0, int(offset or 0))
    if limit is None or int(limit) <= 0:
        return song_urls[o:]
    return song_urls[o:o + int(limit)]


def run_cli_with_split(settings_path: str | None, offset: int, limit: int,
                       per_song_timeout: int, skip_on_timeout: bool, headless: bool):
    save_path, max_items, song_urls = read_initial_settings(settings_path)
    if save_path and song_urls:
        targets = slice_song_range(song_urls, offset, limit)
        if not targets:
            logging.warning('対象曲が0件です（offset/limitを確認）')
            return

        effective_timeout = per_song_timeout
        try:
            if not any(arg.startswith('--per-song-timeout') for arg in sys.argv[2:]):
                effective_timeout = int(CFG.get("PER_SONG_TIME_BUDGET_SEC", per_song_timeout))
        except Exception:
            pass

        function1(save_path, max_items, targets, headless=headless, per_song_timeout=effective_timeout, skip_on_timeout=skip_on_timeout)

        if bool(CFG.get("PARALLEL_ENABLED", False)):
            function2_orchestrator(save_path, targets, headless=headless, per_song_timeout=effective_timeout, skip_on_timeout=skip_on_timeout)
        else:
            function2(save_path, targets, headless=headless, per_song_timeout=effective_timeout, skip_on_timeout=skip_on_timeout)

        logging.info("自動実行が完了しました。")


if __name__ == '__main__':
    try:
        from multiprocessing import freeze_support
        freeze_support()
    except Exception:
        pass

    try:
        kill_chrome_processes()
    except Exception as e:
        logging.warning(f'起動時のChrome/Driver掃除に失敗: {e}')

    service_mode = is_service_like_session()

    if len(sys.argv) > 1 and sys.argv[1] == 'run':
        import argparse
        parser = argparse.ArgumentParser(description='TikTok 明細取得（分割CLI）')
        parser.add_argument('--settings', type=str, default=None)
        parser.add_argument('--offset', type=int, default=0)
        parser.add_argument('--limit', type=int, default=0)
        parser.add_argument('--per-song-timeout', type=int, default=300)
        parser.add_argument('--skip-on-timeout', action='store_true')
        parser.add_argument('--headless', action='store_true')
        args = parser.parse_args(sys.argv[2:])

        headless_flag = args.headless or service_mode
        if service_mode and not args.headless:
            logging.info('サービス/スケジューラからの実行のため、headlessを自動有効化します。')

        run_cli_with_split(settings_path=args.settings, offset=args.offset, limit=args.limit,
                           per_song_timeout=args.per_song_timeout, skip_on_timeout=args.skip_on_timeout, headless=headless_flag)
    else:
        if service_mode:
            logging.info('サービス/スケジューラっぽいセッションのため、CLI(headless)モードで実行します。')
            timeout_cfg = int(CFG.get("PER_SONG_TIME_BUDGET_SEC", 300))
            run_cli_with_split(settings_path=None, offset=0, limit=0, per_song_timeout=timeout_cfg, skip_on_timeout=True, headless=True)
        else:
            create_gui()