# -*- coding: utf-8 -*-
import os
import sys
import csv
import glob
import time
import random
import logging
import argparse
import unicodedata
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
from selenium.common.exceptions import WebDriverException

import config
from modules.excel_utils import (
    read_song_urls,              # 従来: initial_settings.xlsx を読む
    update_difference_entry,
    update_ugc_entry,
    get_row_by_song,
    find_failed_entries,
)
from modules.scraper import get_ugc_count, initialize_driver
from modules.logger import setup_logging
from modules.constants import (
    UGC_SHEET_NAME,
    URL_COLUMN,
    START_ROW,
    MAX_RETRIES,
    RETRY_DELAY,
)

# =============================
#  ユーティリティ
# =============================

CSV_HEADER = ["song", "url", "ugc_count", "timestamp"]
_HLX = re.compile(r'(?i)^\s*=?\s*HYPERLINK\(\s*"([^"]+)"\s*,\s*"([^"]*)"\s*\)\s*$')
_INT_RX = re.compile(r"\d+")

def _parse_int_relaxed(v) -> int | None:
    """'66,600' や 'ＵＧＣ: 8030', '２０６６００' 等から整数を抽出する。
    まずは全角→半角＆桁区切り（, や空白, 全角空白, '_'）を除去して「全桁が数字」なら採用。
    それ以外は従来どおり「先頭の数字塊だけ」を採用（＝既存挙動を保持）。"""
    if v is None:
        return None
    s = unicodedata.normalize("NFKC", str(v)).strip()
    # 桁区切りを除去（値全体が数字なら採用）
    s_all = s.replace(",", "").replace(" ", "").replace("\u3000", "").replace("_", "")
    if s_all.isdigit():
        try:
            return int(s_all)
        except Exception:
            pass
    # 従来の動作：最初に見つかった数字塊のみ
    m = _INT_RX.search(s)
    return int(m.group(0)) if m else None

def _ensure_parent_dir(path: str):
    p = Path(path).parent
    if str(p) and not p.exists():
        p.mkdir(parents=True, exist_ok=True)

# CSV 追記（BOM / 絶対化 / 親作成）
def _append_csv(out_path: str, row: dict):
    p = Path(out_path)
    p = p.expanduser().resolve() if not p.is_absolute() else p
    p.parent.mkdir(parents=True, exist_ok=True)
    file_exists = p.exists()
    with open(p, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(row.keys()))
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)

def _build_shard(items: list, shards: int, shard_index: int) -> list:
    base = max(1, int(shards))
    idx = int(shard_index)
    return [item for i, item in enumerate(items) if i % base == idx]

def _resolve_settings_path(arg_path: str | None) -> Path:
    candidates: list[Path] = []
    if arg_path:
        candidates.append(Path(arg_path))
    candidates.append(Path("initial_settings.xlsx"))
    candidates.append(Path("input") / "initial_settings.xlsx")
    if getattr(config, "EXCEL_FILE_PATH", None):
        candidates.append(Path(config.EXCEL_FILE_PATH))
    for p in candidates:
        if p and Path(p).exists():
            return Path(p).resolve()
    raise FileNotFoundError("initial_settings.xlsx が見つかりません。--settings で明示してください。")

def _resolve_target_path(arg_path: str | None, fallback_settings: Path | None) -> Path:
    if arg_path:
        p = Path(arg_path)
        if p.exists():
            return p.resolve()
        raise FileNotFoundError(f"中央Excelが見つかりません: {arg_path}")
    if fallback_settings:
        return Path(fallback_settings).resolve()
    raise FileNotFoundError("中央Excelのパスが不明です。--target で指定してください。")

def _patch_console_encoding():
    try:
        os.environ.setdefault("PYTHONIOENCODING", "utf-8")
        for h in list(logging.getLogger().handlers):
            if isinstance(h, logging.StreamHandler):
                s = h.stream
                try:
                    s.reconfigure(encoding="utf-8", errors="backslashreplace")
                except Exception:
                    pass
    except Exception:
        pass

def _safe_log_str(s: str) -> str:
    try:
        enc = (sys.stdout.encoding or "utf-8")
        return str(s).encode(enc, "backslashreplace").decode(enc)
    except Exception:
        return str(s).encode("cp932", "backslashreplace").decode("cp932")

def _normalize_name(s: str) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s)).lower()
    t = "".join(ch for ch in t if not ch.isspace())
    return t

def _canonical_url(u: str) -> str:
    if not u:
        return ""
    s = str(u).strip()
    for sep in ["#", "?"]:
        if sep in s:
            s = s.split(sep, 1)[0]
    while s.endswith("/"):
        s = s[:-1]
    return s

def _to_int(v) -> int:
    if isinstance(v, int):
        return v
    return int(str(v))

def _col_index(col) -> int:
    if isinstance(col, int):
        return col
    try:
        return column_index_from_string(str(col))
    except Exception as e:
        raise ValueError(f"URL_COLUMN が不正です（{col}）: {e}")

def _start_row() -> int:
    try:
        return _to_int(START_ROW)
    except Exception:
        return 2

def _extract_hyperlink(cell) -> tuple[str | None, str | None]:
    """
    HYPERLINK を robust に解釈：
      優先1) cell.hyperlink.target
      優先2) =HYPERLINK("url","text")
      優先3) 値が http で始まる素の文字列
    """
    try:
        if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
            url = str(cell.hyperlink.target).strip()
            disp = cell.value if cell.value is not None else ""
            return url, str(disp)
        v = cell.value
        if isinstance(v, str):
            m = _HLX.match(v)
            if m:
                return m.group(1), m.group(2)
            if v.strip().lower().startswith(("http://", "https://")):
                return v.strip(), None
    except Exception:
        pass
    return None, None

def _build_url_index(ugc_sheet) -> dict[str, int]:
    idx: dict[str, int] = {}
    col_url = _col_index(URL_COLUMN)
    srow = _start_row()
    max_row = max(srow, ugc_sheet.max_row or srow)
    for r in range(srow, max_row + 1):
        cell = ugc_sheet.cell(row=r, column=col_url)
        url, _ = _extract_hyperlink(cell)
        if not url:
            val = cell.value
            if isinstance(val, str) and val.strip().lower().startswith(("http://", "https://")):
                url = val.strip()
        if url:
            idx[_canonical_url(url)] = r
    return idx

def _build_name_index(ugc_sheet) -> dict[str, int]:
    srow = _start_row()
    max_row = max(srow, ugc_sheet.max_row or srow)
    max_col = ugc_sheet.max_column or 1
    idx: dict[str, int] = {}
    for r in range(srow, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ugc_sheet.cell(row=r, column=c)
            url, disp = _extract_hyperlink(cell)
            if disp:
                key = _normalize_name(disp)
                if key and key not in idx:
                    idx[key] = r
            v = cell.value
            if isinstance(v, str):
                if _HLX.match(v):
                    continue
                key = _normalize_name(v)
                if key and key not in idx:
                    idx[key] = r
    return idx

def _find_row_by_url_or_name(ugc_sheet, song: str, url: str,
                             url_index: dict[str, int],
                             name_index: dict[str, int]) -> int | None:
    can_url = _canonical_url(url)
    if can_url and can_url in url_index:
        return url_index[can_url]
    key = _normalize_name(song)
    return name_index.get(key)

def _fill_url_if_empty(ugc_sheet, row: int, url: str, url_index: dict[str, int]):
    col_url = _col_index(URL_COLUMN)
    cell = ugc_sheet.cell(row=row, column=col_url)
    has_url = False
    u, _ = _extract_hyperlink(cell)
    if u:
        has_url = True
    else:
        v = cell.value
        if isinstance(v, str) and v.strip().lower().startswith(("http://", "https://")):
            has_url = True
    if not has_url:
        can = _canonical_url(url)
        if can:
            cell.value = can
            url_index[can] = row
            logging.info("[apply] URLを自動補完: row=%d url=%s", row, can)

# =============================
#  新：楽曲マスタ（UGC）読み込み
# =============================

def _normalize_header_text(v: str | None) -> str:
    if v is None:
        return ""
    t = unicodedata.normalize("NFKC", str(v)).strip()
    return t

def _locate_master_header(sheet) -> tuple[int, int, int] | None:
    """
    ヘッダー行（例: '曲名' / 'URL'）を探し、(header_row, col_song, col_url) を返す。
    見つからない場合は None。
    """
    max_row = sheet.max_row or 1
    max_col = sheet.max_column or 1
    for r in range(1, max_row + 1):
        labels = {}
        for c in range(1, max_col + 1):
            txt = _normalize_header_text(sheet.cell(row=r, column=c).value)
            if txt in ("曲名", "楽曲名", "Song", "song"):
                labels["song"] = c
            if txt.upper() == "URL" or txt in ("Url", "url"):
                labels["url"] = c
        if "song" in labels and "url" in labels:
            return (r, labels["song"], labels["url"])
    return None

def _read_songs_from_master(master_xlsx: Path, sheet_name: str) -> list[tuple[str, str]]:
    """
    UGCdate\TikTok_UGC.xlsx の「楽曲マスタ」シートから (song, url) を読み出す。
    - 空セル/空行はスキップ
    - HYPERLINK/素のURLの両方に対応
    - 末尾'/' と '?','#' を除去して正規化
    """
    wb = openpyxl.load_workbook(master_xlsx, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        sh = wb[sheet_name]

        pos = _locate_master_header(sh)
        if not pos:
            # フォールバック: 先頭5行以内にヘッダーが無い場合、A=曲名/B=URL とみなす
            header_row, col_song, col_url = 1, 1, 2
        else:
            header_row, col_song, col_url = pos

        items: list[tuple[str, str]] = []
        for r in range(header_row + 1, (sh.max_row or header_row) + 1):
            song = sh.cell(row=r, column=col_song).value
            url_cell = sh.cell(row=r, column=col_url)
            url, _ = _extract_hyperlink(url_cell)
            if not url:
                v = url_cell.value
                if isinstance(v, str) and v.strip().lower().startswith(("http://", "https://")):
                    url = v.strip()

            if url:
                can = _canonical_url(url)
                s = str(song).strip() if song is not None else f"(row{r})"
                items.append((s, can))
            # URL が無い行はスキップ（CLI仕様：空セルはスキップして処理続行）
        return items
    finally:
        wb.close()

# =============================
#  既存フロー（互換維持）
# =============================

def process_mode(settings_xlsx: Path, driver):
    wb = openpyxl.load_workbook(settings_xlsx)
    try:
        song_url_map = read_song_urls(wb, stop_on_blank=False)
        if not song_url_map:
            logging.warning("曲名とURLのマップを取得できませんでした。処理を中断します。")
            return

        ugc_sheet = wb[UGC_SHEET_NAME]
        url_index = _build_url_index(ugc_sheet)
        name_index = _build_name_index(ugc_sheet)

        for idx, (song, url) in enumerate(song_url_map.items(), start=1):
            logging.info("曲 %d (%s) のUGC数取得を開始します。URL: %s",
                         idx, _safe_log_str(song), url)
            try:
                ugc = get_ugc_count(driver, url, max_retries=MAX_RETRIES, retry_delay=RETRY_DELAY)
            except Exception as e:
                logging.error("UGC取得中にエラー: %s", e)
                ugc = None

            if ugc is None:
                ugc = "取得失敗"

            r = _find_row_by_url_or_name(ugc_sheet, song, url, url_index, name_index)
            if r is None:
                logging.warning("曲 '%s' の行が見つからないためスキップ。", _safe_log_str(song))
                continue

            _fill_url_if_empty(ugc_sheet, r, url, url_index)
            try:
                delta = update_ugc_entry(wb, ugc, r)
                update_difference_entry(wb, delta, r)
            except Exception as e:
                logging.error("Excel更新エラー（%s）: %s", _safe_log_str(song), e)

            try:
                wb.save(settings_xlsx)
            except Exception as e:
                logging.error("Excel保存エラー: %s", e)

            time.sleep(random.uniform(1, 3))
    finally:
        wb.close()

def retry_mode(settings_xlsx: Path, driver):
    base = openpyxl.load_workbook(settings_xlsx)
    try:
        failed = find_failed_entries(base)
        if not failed:
            logging.info("再試行対象なし。")
            return
    finally:
        base.close()

    wb = openpyxl.load_workbook(settings_xlsx)
    try:
        ugc_sheet = wb[UGC_SHEET_NAME]
        url_index = _build_url_index(ugc_sheet)
        name_index = _build_name_index(ugc_sheet)

        for entry in failed:
            row, song, url = entry["row"], entry["song"], entry["url"]
            logging.info("再取得: %s", _safe_log_str(song))
            try:
                ugc = get_ugc_count(driver, url, max_retries=MAX_RETRIES, retry_delay=RETRY_DELAY)
            except Exception as e:
                logging.error("UGC再取得エラー: %s", e)
                ugc = None

            if ugc is None:
                ugc = "取得失敗"

            r = row or _find_row_by_url_or_name(ugc_sheet, song, url, url_index, name_index)
            if not r:
                logging.warning("曲 '%s' の行が見つからずスキップ。", _safe_log_str(song))
                continue

            _fill_url_if_empty(ugc_sheet, r, url, url_index)
            try:
                delta = update_ugc_entry(wb, ugc, r)
                update_difference_entry(wb, delta, r)
            except Exception as e:
                logging.error("Excel更新エラー（%s）: %s", _safe_log_str(song), e)

            try:
                wb.save(settings_xlsx)
            except Exception as e:
                logging.error("Excel保存エラー: %s", e)

            time.sleep(random.uniform(1, 3))
    finally:
        wb.close()

# =============================
#  新：並列 収集/適用
# =============================

def collect_mode(settings_xlsx: Path | None,
                 shards: int, shard_index: int, out_csv: str,
                 profile_dir: str | None = None, timeout: int = 15, retries: int = 3,
                 headless: bool = True, disable_images: bool = True,
                 master_xlsx: Path | None = None, master_sheet: str = "楽曲マスタ"):
    """
    settings_xlsx が与えられれば従来の initial_settings.xlsx を使用。
    master_xlsx が与えられれば UGC の「楽曲マスタ」から読み込む（優先）。
    """
    if master_xlsx:
        items = _read_songs_from_master(master_xlsx, master_sheet)
        logging.info("入力: %s / シート: %s", master_xlsx, master_sheet)
        logging.info("読み込み件数: %d（空セル行はスキップ／空行で停止なし）", len(items))
    else:
        if not settings_xlsx:
            raise FileNotFoundError("collect には --settings もしくは --master-xlsx が必要です。")
        wb = openpyxl.load_workbook(settings_xlsx)
        try:
            song_url_map = read_song_urls(wb, stop_on_blank=False)
            items = list(song_url_map.items())
            logging.info("入力: %s / シート: initial_settings（互換）", settings_xlsx)
            logging.info("読み込み件数: %d（空セル行はスキップ／空行で停止なし）", len(items))
        finally:
            wb.close()

    # 並列分割
    if shards and shards > 1:
        all_cnt = len(items)
        items = _build_shard(items, shards, shard_index)
        logging.info("担当曲数: %d / 総曲数: %d / shards=%d index=%d",
                     len(items), all_cnt, shards, shard_index)
    else:
        logging.info("担当曲数: %d / shards=%d index=%d", len(items), shards, shard_index)

    # WebDriver
    try:
        driver = initialize_driver(profile_dir=profile_dir,
                                   headless=headless,
                                   disable_images=disable_images)
        logging.info("WebDriverを正常に初期化しました。（collect）")
    except WebDriverException as e:
        logging.error("WebDriver初期化に失敗: %s", e)
        sys.exit(1)

    # 収集
    try:
        for idx, (song, url) in enumerate(items, start=1):
            logging.info("[collect] %d) %s", idx, _safe_log_str(song))
            try:
                ugc = get_ugc_count(driver, url, max_retries=retries, retry_delay=timeout)
            except Exception as e:
                logging.error("[collect] 取得中エラー: %s", e)
                ugc = None

            ts = datetime.now().isoformat(timespec="seconds")
            row = {
                "song": song,
                "url": url,
                "ugc_count": ugc if isinstance(ugc, int) else "",  # 失敗は空欄（適用側で自動スキップ）
                "timestamp": ts,
            }
            _append_csv(out_csv, row)
            time.sleep(random.uniform(1, 2))
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        logging.info("WebDriverを終了しました。（collect）")

# ==== apply_mode（恒久対策） ====
def apply_mode(target_xlsx: Path, csv_inputs: list[str]):
    """
    取り込み堅牢版:
      1) 曲名一致（get_row_by_song）最優先
      2) URL インデックスでフォールバック
      3) 正規化名インデックスで最終フォールバック
      4) 取り込み時は utf-8-sig（BOM）で読む
      5) ugc_count は “緩和パース”。非数値はスキップ
      6) URL が空なら補完して以後の一致を安定化
    """
    wb = openpyxl.load_workbook(target_xlsx)
    logging.info("Excelを開きました: %s", target_xlsx)

    applied = 0
    skipped = 0
    url_hit = 0
    name_hit = 0

    try:
        ugc_sheet = wb[UGC_SHEET_NAME]
        url_index = _build_url_index(ugc_sheet)
        name_index = _build_name_index(ugc_sheet)

        for pattern in csv_inputs:
            for csv_path in glob.glob(pattern):
                if not Path(csv_path).is_file():
                    continue

                logging.info("[apply] 取り込み: %s", csv_path)
                with open(csv_path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        song = row.get("song")
                        url = row.get("url")
                        ugc = _parse_int_relaxed(row.get("ugc_count"))

                        if not song or ugc is None:
                            skipped += 1
                            continue

                        # 1) 曲名→行（従来優先度）
                        r = get_row_by_song(ugc_sheet, song, wb)

                        # 2) URL フォールバック（従来と同じ）
                        if r is None and url:
                            can = _canonical_url(url)
                            r = url_index.get(can)
                            if r is not None:
                                url_hit += 1

                        # 3) 正規化名インデックス（新規。既存優先度は崩さない）
                        if r is None:
                            key = _normalize_name(song)
                            r = name_index.get(key)
                            if r is not None:
                                name_hit += 1

                        if r is None:
                            logging.warning("曲名/URL が見つからないためスキップ: %s", _safe_log_str(song))
                            skipped += 1
                            continue

                        # URL 補完（既存ロジック）
                        if url:
                            _fill_url_if_empty(ugc_sheet, r, url, url_index)

                        # 反映（delta=None はログのみ追記し、出力仕様は従来どおり）
                        try:
                            delta = update_ugc_entry(wb, ugc, r)
                            if delta is None:
                                logging.info("[apply] delta=None（前回値が非数値/空など）: row=%d song=%s", r, _safe_log_str(song))
                            update_difference_entry(wb, delta, r)
                            applied += 1
                        except Exception as e:
                            logging.error("[apply] Excel更新エラー（%s）: %s", _safe_log_str(song), e)
                            skipped += 1

        wb.save(target_xlsx)
        logging.info("[apply] 保存完了（反映: %d, スキップ: %d, URL一致: %d, 曲名一致: %d）",
                     applied, skipped, url_hit, name_hit)
    finally:
        wb.close()

# =============================
#  エントリポイント
# =============================

def main():
    parser = argparse.ArgumentParser(description="UGCデータ処理スクリプト（process/retry/collect/apply）")
    parser.add_argument("mode", choices=["process", "retry", "collect", "apply"], help="実行モード")

    # 旧・互換
    parser.add_argument("--settings", default=None, help="設定Excel（曲リスト）。未指定時は自動解決")
    parser.add_argument("--target", default=None, help="中央Excel（UGC集計）。未指定時は --settings と同一を使用")

    # 新・UGC 楽曲マスタ（collect 用）
    parser.add_argument("--master-xlsx", default=None, help="UGC側の楽曲マスタ Excel 例: UGCdate\\TikTok_UGC.xlsx")
    parser.add_argument("--master-sheet", default="楽曲マスタ", help="マスタのシート名（既定: 楽曲マスタ）")

    # collect
    parser.add_argument("--shards", type=int, default=1, help="総ワーカー数")
    parser.add_argument("--shard-index", type=int, default=0, help="自ワーカーのインデックス（0開始）")
    parser.add_argument("--out", dest="out_csv", default="runs/ugc.csv", help="収集CSVの出力先")
    parser.add_argument("--profile-dir", default=None, help="Chromeユーザーデータディレクトリ")
    parser.add_argument("--timeout", type=int, default=RETRY_DELAY, help="リトライ間隔の基準秒（collect時）")
    parser.add_argument("--retries", type=int, default=MAX_RETRIES, help="リトライ回数（collect時）")
    parser.add_argument("--no-headless", action="store_true", help="ヘッドレス無効（デバッグ用）")
    parser.add_argument("--enable-images", action="store_true", help="画像読み込みを有効化")

    # apply
    parser.add_argument("--in", dest="csv_inputs", nargs="*", default=[], help="適用対象CSV（複数/ワイルドカード可）")

    args = parser.parse_args()
    setup_logging()
    _patch_console_encoding()

    # 設定ブック解決（collect/process/retryで使用）
    try:
        settings_path = _resolve_settings_path(args.settings) if args.settings else None
    except FileNotFoundError as e:
        if args.mode in ("process", "retry") and not args.master_xlsx:
            logging.error(str(e))
            sys.exit(1)
        settings_path = None  # collect で master 指定があれば不要

    if args.mode == "collect":
        # master 指定があれば優先
        master_path = Path(args.master_xlsx).resolve() if args.master_xlsx else None
        if master_path and not master_path.exists():
            logging.error("指定の --master-xlsx が見つかりません: %s", master_path)
            sys.exit(1)

        collect_mode(
            settings_xlsx=settings_path,
            shards=args.shards,
            shard_index=args.shard_index,
            out_csv=args.out_csv,
            profile_dir=args.profile_dir,
            timeout=args.timeout,
            retries=args.retries,
            headless=(not args.no_headless),
            disable_images=(not args.enable_images),
            master_xlsx=master_path,
            master_sheet=args.master_sheet,
        )
        logging.info("スクリプトの実行を終了します。")
        return

    if args.mode == "apply":
        try:
            target_path = _resolve_target_path(args.target, settings_path)
        except FileNotFoundError as e:
            logging.error(str(e))
            sys.exit(1)

        if not args.csv_inputs:
            logging.error("apply には --in でCSVパターンを指定してください。例: --in runs\\w1\\ugc_*.csv runs\\w2\\ugc_*.csv")
            sys.exit(2)
        apply_mode(target_xlsx=target_path, csv_inputs=args.csv_inputs)
        logging.info("スクリプトの実行を終了します。")
        return

    # 既存の process / retry（互換運用）
    try:
        driver = initialize_driver()
        logging.info("WebDriverを正常に初期化しました。")
    except WebDriverException as e:
        logging.error("WebDriver初期化に失敗: %s", e)
        sys.exit(1)

    try:
        if args.mode == "process":
            if not settings_path:
                logging.error("process モードには --settings が必要です。")
                sys.exit(1)
            process_mode(settings_xlsx=settings_path, driver=driver)
        else:
            if not settings_path:
                logging.error("retry モードには --settings が必要です。")
                sys.exit(1)
            retry_mode(settings_xlsx=settings_path, driver=driver)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        logging.info("スクリプトの実行を終了します。")

if __name__ == "__main__":
    main()
