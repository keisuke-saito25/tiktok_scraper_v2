import openpyxl
import logging
import config
import re
from datetime import datetime, timedelta
from modules.constants import (
    UGC_SHEET_NAME,
    DIFFERENCE_SHEET_NAME,
    URL_COLUMN,
    HEADER_ROW,
    START_ROW,
    DATE_FORMAT,
    ALERT_CELL,
    DELTA_COLUMN,
    RATIO_COLUMN,
    MIN_COL_FOR_COMPARISON,
    ALERT_FILL_COLOR,
    ALERT_FONT_COLOR,
    DEFAULT_FILL_TYPE,
    DEFAULT_FONT_COLOR
)
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def read_song_urls(workbook):
    """
    ExcelファイルのA列とB列から曲名とURLのマップを読み取る関数
    """
    try:
        sheet = workbook.active
        song_url_map = {}
        song_column = 'A'
        url_column = 'B'
        header_row = HEADER_ROW

        current_row = START_ROW
        while True:
            song_name = sheet[f'{song_column}{current_row}'].value
            url = sheet[f'{url_column}{current_row}'].value
            if song_name is None and url is None:
                break  # 両方が空の場合は終了
            if song_name and url:
                song_url_map[song_name] = url
            else:
                logging.warning("行 %d に不完全なデータがあります。曲名: %s, URL: %s", current_row, song_name, url)
            current_row += 1

        return song_url_map

    except Exception as e:
        logging.error("曲名とURLの読み取り中にエラーが発生しました: %s", e)
        return {}

def read_alert_value(workbook):
    """
    B1セルからアラートの基準値を読み込む関数
    """
    try:
        sheet = workbook.active
        b1_value = sheet[ALERT_CELL].value

        if b1_value is None:
            logging.error("B1セルが空です。アラート基準値を設定してください。")
            return None

        try:
            alert_value = float(b1_value)
            logging.info("アラート基準値を %0.2f%% として読み込みました。", alert_value)
            return alert_value
        except ValueError:
            logging.error("B1セルの値が数値ではありません: %s", b1_value)
            return None

    except Exception as e:
        logging.error("アラート基準値の読み取り中にエラーが発生しました: %s", e)
        return None

def read_urls(workbook):
    """
    ExcelファイルのB4セルから下にあるURLを全て読み取る関数
    """
    try:
        sheet = workbook.active
        urls = []
        column = URL_COLUMN
        header = sheet[f'{column}{HEADER_ROW}'].value

        if header != 'URL':
            logging.warning("期待するヘッダー 'URL' がセル %s%d に見つかりません。実際の値: %s", column, HEADER_ROW, header)
            return urls
        
        current_row = HEADER_ROW + 1
        while True:
            cell_value = sheet[f'{column}{current_row}'].value
            if cell_value is None:
                break  # 空セルに達したら終了
            urls.append(cell_value)
            current_row += 1
            
        return urls
        
    except Exception as e:
        logging.error("URLの読み取り中にエラーが発生しました: %s", e)
        return []

def get_sheet(workbook, sheet_name):
    """
    指定されたシート名のシートを取得する関数
    """
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    else:
        logging.error("シート '%s' が存在しません。", sheet_name)
        raise ValueError(f"シート '{sheet_name}' が存在しません。")

def get_or_create_today_column(sheet, header_row=4):
    """
    当日日付の列を取得する。存在しなければ最終列の次に追加
    """
    # シートのクリーニング
    clean_sheet(sheet, header_row)  

    today_str = datetime.now().strftime(DATE_FORMAT)

    # ヘッダー行のセルを取得
    header_cells = sheet[header_row]
    header_values = [cell.value for cell in header_cells]

    # 当日日付がすでに存在するか確認
    try:
        col = header_values.index(today_str) + 1
        return col, False
    except ValueError:
        last_col_with_data = get_last_column_with_data(header_values)
        new_col = last_col_with_data + 1
        sheet.cell(row=header_row, column=new_col).value = today_str
        return new_col, True  # 新規作成

def get_last_column_with_data(header_values):
    """
    ヘッダー行の値から、最後にデータが存在する列番号を返す
    """
    last_col = 0
    for idx, value in enumerate(header_values, start=1):
        if value is not None:
            last_col = idx
    return last_col


def clean_sheet(sheet, header_row=4):
    """
    ヘッダー行以降で不必要な最終列を削除する関数
    """
    for col in range(sheet.max_column, 0, -1):
        # ヘッダー行のセルを基準にチェック
        cell = sheet.cell(row=header_row, column=col)
        if cell.value is None and not any(sheet.cell(row=row, column=col).value for row in range(header_row+1, sheet.max_row+1)):
            sheet.delete_cols(col)
        else:
            break  # データが存在する列が見つかったら終了

def apply_style(cell, fill, font):
    """
    セルにスタイルを適用する関数
    """
    cell.fill = fill
    cell.font = font

def check_and_apply_alert(cell_delta, cell_ratio, ratio, alert_value, styles, row):
    """
    アラート条件をチェックし、セルにスタイルを適用またはリセットする関数
    """
    alert_fill, alert_font, default_fill, default_font = styles
    if alert_value is not None and ratio >= alert_value:
        logging.debug("アラート条件を満たしました。行: %d, ratio: %.2f%% >= %.2f%%", row, ratio, alert_value)
        apply_style(cell_delta, alert_fill, alert_font)
        apply_style(cell_ratio, alert_fill, alert_font)
    else:
        logging.debug("アラート条件を満たしていません。行: %d, ratio: %.2f%% < %.2f%%", row, ratio, alert_value)
        apply_style(cell_delta, default_fill, default_font)
        apply_style(cell_ratio, default_fill, default_font)


def get_row_by_song(sheet, song_name, workbook):
    """
    曲名から対応する行番号を取得する関数。
    セルが数式の場合、参照先の値を取得して比較します。
    """
    for row in range(START_ROW, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)  # A列
        cell_value = cell.value

        # セルが数式かどうかをチェック
        if isinstance(cell_value, str) and cell_value.startswith('='):
            referenced_sheet_name, referenced_cell = parse_formula(cell_value)
            if referenced_sheet_name and referenced_cell:
                # 参照先シートが存在するか確認
                if referenced_sheet_name in workbook.sheetnames:
                    referenced_sheet = workbook[referenced_sheet_name]
                    referenced_value = referenced_sheet[referenced_cell].value
                    if referenced_value == song_name:
                        return row
                else:
                    logging.warning("参照先シート '%s' が存在しません。", referenced_sheet_name)
        else:
            # セルが数式でない場合の通常の比較
            if cell_value == song_name:
                return row

    logging.warning("曲名 '%s' に対応する行が見つかりません。", song_name)
    return None

def update_ugc_entry(workbook, ugc, row):
    """
    UGCシートに単一のUGC数を書き込み、deltaとratioを更新する関数
    rowはシート内の行番号
    """
    try:
        ugc_sheet = get_sheet(workbook, UGC_SHEET_NAME)
    except ValueError as e:
        logging.error(e)
        return

    # アラート基準値を読み込む
    alert_value = read_alert_value(workbook)
    if alert_value is None:
        logging.warning("アラート基準値が設定されていません。")

    today_col, is_new = get_or_create_today_column(ugc_sheet)
    if is_new:
        logging.info("UGCシートに新しい日付の列を追加しました。列番号: %d", today_col)

    if isinstance(ugc, (int, float)):
        # 当日のUGC数を書き込む
        ugc_sheet.cell(row=row, column=today_col).value = ugc

        # 前日との比較
        if today_col <= MIN_COL_FOR_COMPARISON:
            delta = None
            ratio = None
            logging.debug("行 %d の前日データが不足しています。", row)
        else:
            prev_col = today_col - 1
            prev_ugc = ugc_sheet.cell(row=row, column=prev_col).value
            if prev_ugc is None or not isinstance(prev_ugc, (int, float)):
                logging.warning("行 %d の前日のUGC数が無効です。", row)
                delta = None
                ratio = None
            else:
                delta = ugc - prev_ugc
                ratio = (delta / prev_ugc) * 100 if prev_ugc != 0 else 0

        logging.debug("行 %d: delta=%s, ratio=%s%%, alert_value=%s%%", row, delta, ratio, alert_value)

        # 増減数 (B列) と増減率 (C列) を更新
        cell_delta = ugc_sheet.cell(row=row, column=DELTA_COLUMN)
        cell_ratio = ugc_sheet.cell(row=row, column=RATIO_COLUMN)
        
        if delta is not None:
            cell_delta.value = delta
        else:
            cell_delta.value = None  # 空白にする

        if ratio is not None:
            cell_ratio.value = f"{ratio:.2f}%"
        else:
            cell_ratio.value = None  # 空白にする

        if delta is not None and ratio is not None:
            # スタイルの定義
            alert_fill = PatternFill(start_color=ALERT_FILL_COLOR, end_color=ALERT_FILL_COLOR, fill_type="solid")
            alert_font = Font(bold=True, color=ALERT_FONT_COLOR)
            default_fill = PatternFill(fill_type=DEFAULT_FILL_TYPE)
            default_font = Font(bold=False, color=DEFAULT_FONT_COLOR)
            styles = (alert_fill, alert_font, default_fill, default_font)

            # アラートチェック
            check_and_apply_alert(cell_delta, cell_ratio, ratio, alert_value, styles, row)
        else:
            # スタイルをデフォルトにリセット
            default_fill = PatternFill(fill_type=DEFAULT_FILL_TYPE)
            default_font = Font(bold=False, color=DEFAULT_FONT_COLOR)
            apply_style(cell_delta, default_fill, default_font)
            apply_style(cell_ratio, default_fill, default_font)
        
        return delta

    else:
        # UGC取得失敗の場合
        ugc_sheet.cell(row=row, column=today_col).value = ugc  # "取得失敗" を設定

        # 増減数 (B列) と増減率 (C列) を空白にする
        cell_delta = ugc_sheet.cell(row=row, column=DELTA_COLUMN)
        cell_ratio = ugc_sheet.cell(row=row, column=RATIO_COLUMN)
        cell_delta.value = None
        cell_ratio.value = None

        # スタイルをデフォルトにリセット
        default_fill = PatternFill(fill_type=DEFAULT_FILL_TYPE)
        default_font = Font(bold=False, color=DEFAULT_FONT_COLOR)
        apply_style(cell_delta, default_fill, default_font)
        apply_style(cell_ratio, default_fill, default_font)
        
        return None  # delta_count は None

def update_difference_entry(workbook, delta, row):
    """
    増減シートに増減数を書き込む関数
    rowはシート内の行番号
    """
    try:
        difference_sheet = get_sheet(workbook, DIFFERENCE_SHEET_NAME)
    except ValueError as e:
        logging.error(e)
        return

    today_str = datetime.now().strftime(DATE_FORMAT)
    today_col, is_new = get_or_create_today_column(difference_sheet)
    if is_new: 
        difference_sheet.cell(row=HEADER_ROW, column=today_col).value = today_str
        logging.info("増減シートに新しい日付の列を追加しました: %s", today_str)

    song_name = difference_sheet.cell(row=row, column=1).value
    if not song_name:
        logging.warning("行 %d に曲名が存在しません。", row)
        return
    
    if delta is not None:
        difference_sheet.cell(row=row, column=today_col).value = delta
    else:
        difference_sheet.cell(row=row, column=today_col).value = None  # 空白にする

def find_failed_entries(workbook):
    """
    UGCシートから最新の日付列に '取得失敗' とマークされている行を取得する関数
    """
    ugc_sheet = workbook[UGC_SHEET_NAME]
    
    # ヘッダー行から最新の日付列を特定
    header_cells = ugc_sheet[HEADER_ROW]
    date_columns = []
    for cell in header_cells:
        if cell.value and isinstance(cell.value, str):
            try:
                datetime.strptime(cell.value, DATE_FORMAT)
                date_columns.append(cell.column)
            except ValueError:
                continue  # 日付でない場合はスキップ
    
    if not date_columns:
        logging.error("UGCシートに有効な日付列が見つかりません。")
        return []
    
    latest_date_col = max(date_columns)  # 最も右側の列が最新の日付
    latest_date_str = ugc_sheet.cell(row=HEADER_ROW, column=latest_date_col).value
    logging.info("最新の日付列を特定しました: %s (列番号: %d)", latest_date_str, latest_date_col)
    
    failed_entries = []
    
    # URLシートの読み込み
    try:
        url_sheet = get_sheet(workbook, '楽曲マスタ') 
    except ValueError as e:
        logging.error(e)
        return []
    
    for row in range(START_ROW, ugc_sheet.max_row + 1):
        status = ugc_sheet.cell(row=row, column=latest_date_col).value
        if status == "取得失敗":
            cell = ugc_sheet.cell(row=row, column=1)  # A列
            song_name = None

            if isinstance(cell.value, str) and cell.value.startswith('='):
                # 数式の場合、参照先の値を取得
                referenced_sheet_name, referenced_cell = parse_formula(cell.value)
                if referenced_sheet_name and referenced_cell:
                    if referenced_sheet_name in workbook.sheetnames:
                        referenced_sheet = workbook[referenced_sheet_name]
                        song_name = referenced_sheet[referenced_cell].value
                        logging.debug("数式から取得した曲名: %s", song_name)
                    else:
                        logging.warning("参照先シート '%s' が存在しません。", referenced_sheet_name)
            else:
                # 通常の値の場合
                song_name = cell.value

            if not song_name:
                logging.warning("行 %d の曲名が取得できませんでした。", row)
                continue

            # URLシートから曲名に対応するURLを検索
            url = find_url_by_song_name(url_sheet, song_name)
            if url:
                failed_entries.append({
                    'row': row,
                    'song': song_name,
                    'url': url
                })
                logging.info("再試行対象 - 行: %d, 曲名: %s, URL: %s", row, song_name, url)
            else:
                logging.warning("曲名 '%s' に対応するURLが見つかりません。", song_name)
    
    return failed_entries

def parse_formula(formula):
    """
    数式から参照先シート名とセルアドレスを抽出する関数
    例: "= '楽曲マスタ'!A6" -> ('楽曲マスタ', 'A6')
    """
    import re
    pattern = re.compile(r"='?([^'!]+)'?!([A-Z]+[0-9]+)")
    match = pattern.match(formula)
    if match:
        referenced_sheet = match.group(1)
        referenced_cell = match.group(2)
        return referenced_sheet, referenced_cell
    else:
        logging.warning("予期しない数式形式: %s", formula)
        return None, None

def find_url_by_song_name(url_sheet, song_name):
    """
    URLシートから曲名に対応するURLを検索する関数
    """
    for row in range(START_ROW, url_sheet.max_row + 1):
        cell_song = url_sheet.cell(row=row, column=1).value  # A列
        cell_url = url_sheet.cell(row=row, column=2).value   # B列
        if cell_song == song_name:
            return cell_url
    return None
